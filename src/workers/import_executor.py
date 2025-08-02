"""Executor for import jobs with streaming and memory efficiency."""

import os
import json
import hashlib
import asyncio
import csv
import zipfile
from typing import Dict, Any, List, Optional, AsyncIterator, Tuple
from datetime import datetime
from uuid import UUID
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed
import io
import tempfile

import aiofiles
import aiofiles.os
import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl.utils.exceptions import InvalidFileException
import xxhash
import psycopg

from src.workers.job_worker import JobExecutor
from src.infrastructure.postgres.database import DatabasePool
from src.infrastructure.postgres.event_store import PostgresEventStore
from src.core.events.publisher import JobStartedEvent, JobCompletedEvent, JobFailedEvent
from src.core.events.registry import InMemoryEventBus
from src.infrastructure.config import get_settings


class ImportJobExecutor(JobExecutor):
    """Executes import jobs with optimized streaming pipeline."""
    
    def __init__(self):
        self.settings = get_settings()
        self.batch_size = self.settings.import_batch_size
        # Use xxhash for much faster hashing (50-75x faster than SHA256)
        self.use_xxhash = True
        self.xxhash_seed = 0
        # Parallel processing settings
        self.parallel_workers = self.settings.import_parallel_workers
        self.parallel_threshold_mb = self.settings.import_parallel_threshold_mb
        # Construct database URL for worker processes
        self.db_url = f"postgresql://{self.settings.POSTGRESQL_USER}:{self.settings.POSTGRESQL_PASSWORD}@{self.settings.POSTGRESQL_HOST}:{self.settings.POSTGRESQL_PORT}/{self.settings.POSTGRESQL_DATABASE}"
    
    async def execute(self, job_id: str, parameters: Dict[str, Any], db_pool: DatabasePool) -> Dict[str, Any]:
        """Execute import job with streaming pipeline."""
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f"Import job {job_id} - Starting optimized streaming import")
        
        # Create event bus and store for publishing events
        event_store = PostgresEventStore(db_pool)
        event_bus = InMemoryEventBus()
        event_bus.set_event_store(event_store)
        
        # Handle case where parameters come as string
        if isinstance(parameters, str):
            parameters = json.loads(parameters)
        
        logger.info(f"Import job {job_id} - Parameters: {parameters}")
        
        # Extract parameters
        temp_file_path = parameters.get('temp_file_path', parameters.get('file_path'))
        filename = parameters.get('filename', parameters.get('file_name'))
        
        # Ensure temp_file_path exists before proceeding
        if not await aiofiles.os.path.exists(temp_file_path):
            raise FileNotFoundError(f"Temporary import file not found at path: {temp_file_path}")
        
        # Get job details from database
        async with db_pool.acquire() as conn:
            job = await conn.fetchrow(
                "SELECT source_commit_id, dataset_id, user_id FROM dsa_jobs.analysis_runs WHERE id = $1",
                UUID(job_id)
            )
            if not job:
                raise ValueError(f"Job with ID {job_id} not found.")
            parent_commit_id = job['source_commit_id']
            dataset_id = parameters.get('dataset_id', job['dataset_id'])
            user_id = parameters.get('user_id', job['user_id'])
        
        try:
            # Publish job started event
            await event_bus.publish(JobStartedEvent(
                job_id=job_id,
                job_type='import',
                dataset_id=dataset_id,
                user_id=user_id
            ))
            
            # Update job progress - starting
            await self._update_job_progress(job_id, {"status": "Starting import", "percentage": 0}, db_pool)
            
            # Create commit first
            commit_id = await self._create_commit(
                db_pool, dataset_id, parent_commit_id, user_id, parameters['commit_message']
            )
            
            # Determine file type and process with appropriate method
            logger.info(f"Import job {job_id} - Processing file '{filename}' with extension '{os.path.splitext(filename)[1].lower()}'")
            file_ext = os.path.splitext(filename)[1].lower()
            
            if file_ext == '.csv':
                total_rows_processed = await self._process_csv_file(
                    commit_id, temp_file_path, job_id, db_pool
                )
            elif file_ext == '.xlsx':
                total_rows_processed = await self._process_excel_file(
                    commit_id, temp_file_path, job_id, db_pool
                )
            elif file_ext == '.parquet':
                total_rows_processed = await self._process_parquet_file(
                    commit_id, temp_file_path, job_id, db_pool
                )
            else:
                raise ValueError(f"Unsupported file format: {file_ext}. Supported formats are: .csv, .xlsx, .parquet")
            
            # Update ref to point to new commit
            await self._update_ref(db_pool, dataset_id, parameters.get('target_ref', 'main'), commit_id)
            
            # Run post-import maintenance tasks
            await self._run_post_import_maintenance(commit_id, job_id, db_pool)
            
            # Update final job status
            await self._update_job_progress(
                job_id, 
                {"status": "Completed", "percentage": 100, "rows_processed": total_rows_processed}, 
                db_pool
            )
            
            # Publish job completed event
            await event_bus.publish(JobCompletedEvent(
                job_id=job_id,
                status='completed',
                dataset_id=dataset_id,
                result={
                    "commit_id": commit_id,
                    "rows_imported": total_rows_processed
                }
            ))
            
            return {
                "commit_id": commit_id,
                "rows_imported": total_rows_processed,
                "message": f"Successfully imported {total_rows_processed:,} rows"
            }
            
        except Exception as e:
            # Mark job as failed
            import traceback
            logger.error(f"Import job {job_id} failed: {e}\n{traceback.format_exc()}")
            
            # Publish job failed event
            await event_bus.publish(JobFailedEvent(
                job_id=job_id,
                error_message=str(e),
                dataset_id=dataset_id
            ))
            
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "UPDATE dsa_jobs.analysis_runs SET status = 'failed', error_message = $2 WHERE id = $1",
                    UUID(job_id), str(e)
                )
            raise
        finally:
            # IMPROVEMENT: Use a finally block to guarantee cleanup
            if await aiofiles.os.path.exists(temp_file_path):
                await aiofiles.os.remove(temp_file_path)
                logger.info(f"Import job {job_id} - Cleaned up temp file: {temp_file_path}")
    
    def _calculate_hash(self, data: str) -> str:
        """Calculate hash of data using xxHash (50-75x faster than SHA256) or SHA256."""
        if self.use_xxhash:
            return xxhash.xxh64(data.encode(), seed=self.xxhash_seed).hexdigest()
        else:
            return hashlib.sha256(data.encode()).hexdigest()
    
    async def _process_and_commit_batch(
        self, 
        batch: List[Tuple[int, Dict]], 
        sheet_name: str, 
        commit_id: str, 
        db_pool: DatabasePool
    ) -> None:
        """Process and immediately commit a batch to the database."""
        if not batch:
            return

        async with db_pool.acquire() as conn:
            async with conn.transaction():
                # Using a generator expression to build copy_data is more memory-efficient
                def prepare_copy_data():
                    for line_num, row_data in batch:
                        data_json = json.dumps(row_data, sort_keys=True, separators=(',', ':'))
                        data_hash = self._calculate_hash(data_json)
                        logical_row_id = f"{sheet_name}:{line_num}"  # Use line number instead of hash
                        
                        # Store only the raw row data, metadata is in logical_row_id
                        yield (logical_row_id, data_hash, data_json)

                # Use COPY directly with pg_temp schema
                await conn.copy_records_to_table(
                    'dsa_core_temp_import_rows',
                    records=list(prepare_copy_data()),
                    columns=['logical_row_id', 'row_hash', 'data'],
                    schema_name='pg_temp'
                )
                
                await conn.execute("""
                    INSERT INTO dsa_core.rows (row_hash, data)
                    SELECT t.row_hash, t.data::jsonb FROM pg_temp.dsa_core_temp_import_rows t
                    ON CONFLICT (row_hash) DO NOTHING
                """)
                
                await conn.execute("""
                    INSERT INTO dsa_core.commit_rows (commit_id, logical_row_id, row_hash)
                    SELECT $1, t.logical_row_id, t.row_hash FROM pg_temp.dsa_core_temp_import_rows t
                """, commit_id)
    
    async def _process_csv_file(
        self, 
        commit_id: str,
        file_path: str, 
        job_id: str, 
        db_pool: DatabasePool
    ) -> int:
        """Process CSV file with true streaming, avoiding loading the whole file into memory."""
        # Check if we should use parallel processing
        if await self._should_use_parallel_processing(file_path):
            return await self._process_csv_file_parallel(commit_id, file_path, job_id, db_pool)
        
        # Otherwise, use sequential processing for smaller files
        sheet_name = 'primary'
        total_rows_processed = 0
        
        try:
            async with aiofiles.open(file_path, mode='r', encoding='utf-8-sig', newline='') as afp:
                # Get headers first
                header_line = await afp.readline()
                headers = next(csv.reader([header_line]))
                
                batch = []
                line_number = 2  # Start at line 2 (after header)
                
                # Process line by line without loading the whole file
                async for line in afp:
                    # Create a dict from headers and the current line's values
                    row_data = dict(zip(headers, next(csv.reader([line]))))
                    batch.append((line_number, row_data))
                    line_number += 1

                    if len(batch) >= self.batch_size:
                        await self._process_and_commit_batch(batch, sheet_name, commit_id, db_pool)
                        total_rows_processed += len(batch)
                        batch = []
                        
                        await self._update_job_progress(
                            job_id,
                            {"status": f"Processing: {total_rows_processed:,} rows", "rows_processed": total_rows_processed},
                            db_pool
                        )
                
                if batch:
                    await self._process_and_commit_batch(batch, sheet_name, commit_id, db_pool)
                    total_rows_processed += len(batch)

        except (UnicodeDecodeError, csv.Error) as e:
            raise ValueError(f"Failed to parse CSV file. Ensure it is a valid UTF-8 encoded CSV. Error: {e}")

        return total_rows_processed
    
    
    async def _process_excel_file(
        self, 
        commit_id: str,
        file_path: str, 
        job_id: str, 
        db_pool: DatabasePool
    ) -> int:
        """Process Excel (.xlsx) file using a thread to avoid blocking the event loop."""
        
        def blocking_excel_read():
            """Synchronous function to run in a separate thread."""
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            except (InvalidFileException, zipfile.BadZipFile) as e:
                raise ValueError(f"Failed to open Excel file. It may be corrupt or an unsupported format (e.g., .xls instead of .xlsx). Error: {e}")

            # This generator yields batches to the async part of the code
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows_iter = sheet.iter_rows()
                
                try:
                    headers = [cell.value for cell in next(rows_iter)]
                except StopIteration:
                    continue  # Skip empty sheet
                
                batch = []
                sheet_row_count = 2  # Start at row 2 (after header)
                
                for row_idx, row in enumerate(rows_iter, start=2):
                    row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                    batch.append((row_idx, row_data))
                    
                    if len(batch) >= self.batch_size:
                        yield batch, sheet_name
                        batch = []
                
                if batch:
                    yield batch, sheet_name
                    # No need to update sheet_row_count here since we're done with the sheet
            
            wb.close()

        total_rows_processed = 0
        # Run the blocking generator in a thread pool
        loop = asyncio.get_running_loop()
        blocking_iterator = await loop.run_in_executor(None, blocking_excel_read)

        for batch, sheet_name in blocking_iterator:
            await self._process_and_commit_batch(batch, sheet_name, commit_id, db_pool)
            total_rows_processed += len(batch)
            
            await self._update_job_progress(
                job_id,
                {"status": f"Processing {sheet_name}: {total_rows_processed:,} total rows", "rows_processed": total_rows_processed},
                db_pool
            )

        return total_rows_processed
    
    async def _process_parquet_file(
        self, 
        commit_id: str,
        file_path: str, 
        job_id: str, 
        db_pool: DatabasePool
    ) -> int:
        """Process Parquet file using a thread to avoid blocking CPU-bound conversion."""
        sheet_name = 'primary'
        total_rows_processed = 0

        def blocking_parquet_read():
            """Process Parquet file in a thread using Polars for high performance."""
            import polars as pl
            
            row_number = 2  # Start at row 2 (treating as if there's a header)
            
            # Polars can work directly with Arrow batches from PyArrow
            parquet_file = pq.ParquetFile(file_path)
            
            for batch in parquet_file.iter_batches(batch_size=self.batch_size):
                # Convert arrow batch directly to Polars DF (zero-copy when possible)
                df_chunk = pl.from_arrow(batch)
                
                # iter_rows with named=True creates dicts one-by-one (memory efficient)
                batch_with_nums = [
                    (row_number + i, record) 
                    for i, record in enumerate(df_chunk.iter_rows(named=True))
                ]
                row_number += len(batch_with_nums)
                yield batch_with_nums

        loop = asyncio.get_running_loop()
        blocking_iterator = await loop.run_in_executor(None, blocking_parquet_read)
        
        for batch in blocking_iterator:
            await self._process_and_commit_batch(batch, sheet_name, commit_id, db_pool)
            total_rows_processed += len(batch)
            
            await self._update_job_progress(
                job_id,
                {"status": f"Processing: {total_rows_processed:,} rows", "rows_processed": total_rows_processed},
                db_pool
            )
        
        return total_rows_processed
    
    async def _create_commit(
        self, db_pool: DatabasePool, dataset_id: int, parent_commit_id: Optional[str],
        user_id: int, message: str
    ) -> str:
        """Create a new commit."""
        import uuid
        # Use standard UUID v4 for guaranteed uniqueness
        commit_id = uuid.uuid4().hex
        async with db_pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO dsa_core.commits (commit_id, dataset_id, parent_commit_id, author_id, message, authored_at, committed_at) "
                "VALUES ($1, $2, $3, $4, $5, NOW(), NOW())",
                commit_id, dataset_id, parent_commit_id, user_id, message
            )
        return commit_id
    
    async def _update_ref(
        self, db_pool: DatabasePool, dataset_id: int, ref_name: str, commit_id: str
    ) -> None:
        """Update reference to point to new commit."""
        async with db_pool.acquire() as conn:
            # Using ON CONFLICT makes this operation idempotent - it can create or update the ref
            await conn.execute("""
                INSERT INTO dsa_core.refs (dataset_id, name, commit_id)
                VALUES ($1, $2, $3)
                ON CONFLICT (dataset_id, name) DO UPDATE 
                SET commit_id = $3
            """, dataset_id, ref_name, commit_id)
    
    async def _update_job_progress(
        self, job_id: str, progress_info: Dict[str, Any], db_pool: DatabasePool
    ) -> None:
        """Update job progress in analysis_runs table."""
        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE dsa_jobs.analysis_runs "
                "SET run_parameters = jsonb_set(run_parameters, '{progress}', $1::jsonb, true) "
                "WHERE id = $2",
                json.dumps(progress_info), UUID(job_id)
            )
    
    async def _analyze_imported_tables(self, commit_id: str, db_pool: DatabasePool) -> None:
        """Analyze imported tables and store schema/statistics."""
        async with db_pool.acquire() as conn:
            # Get all unique table keys for this commit
            table_keys = await conn.fetch("""
                SELECT DISTINCT split_part(logical_row_id, ':', 1) as table_key
                FROM dsa_core.commit_rows
                WHERE commit_id = $1
            """, commit_id)
            
            for record in table_keys:
                table_key = record['table_key']
                
                # Get sample rows to analyze schema
                sample_rows = await conn.fetch("""
                    SELECT r.data
                    FROM dsa_core.commit_rows cr
                    JOIN dsa_core.rows r ON cr.row_hash = r.row_hash
                    WHERE cr.commit_id = $1 
                    AND cr.logical_row_id LIKE $2 || ':%'
                    LIMIT 100
                """, commit_id, table_key)
                
                if not sample_rows:
                    continue
                
                # Analyze schema from sample
                all_columns = set()
                column_types = {}
                null_counts = {}
                unique_values = {}
                
                for row in sample_rows:
                    # The data column in dsa_core.rows should be jsonb, so no need for json.loads
                    row_data = row['data']
                    
                    # Skip if row_data is not a dict
                    if not isinstance(row_data, dict):
                        continue
                    
                    for col, value in row_data.items():
                        all_columns.add(col)
                        
                        # Track nulls
                        if col not in null_counts:
                            null_counts[col] = 0
                        if value is None:
                            null_counts[col] += 1
                        
                        # Track unique values
                        if col not in unique_values:
                            unique_values[col] = set()
                        if value is not None and len(unique_values[col]) < 100:
                            unique_values[col].add(str(value))
                        
                        # Improved type inference - check all values, not just the first
                        if value is not None:
                            current_type = column_types.get(col, None)
                            value_type = None
                            
                            if isinstance(value, bool):
                                value_type = 'boolean'
                            elif isinstance(value, int):
                                value_type = 'integer'
                            elif isinstance(value, float):
                                value_type = 'float'
                            else:
                                value_type = 'string'
                            
                            # Type promotion logic
                            if current_type is None:
                                column_types[col] = value_type
                            elif current_type != value_type:
                                # Promote types: boolean < integer < float < string
                                type_hierarchy = {'boolean': 0, 'integer': 1, 'float': 2, 'string': 3}
                                if type_hierarchy.get(value_type, 3) > type_hierarchy.get(current_type, 3):
                                    column_types[col] = value_type
                
                # Get total row count
                count_result = await conn.fetchrow("""
                    SELECT COUNT(*) as total
                    FROM dsa_core.commit_rows
                    WHERE commit_id = $1 AND logical_row_id LIKE $2 || ':%'
                """, commit_id, table_key)
                total_rows = count_result['total']
                
                # Build analysis structure
                analysis = {
                    'total_rows': total_rows,
                    'column_types': column_types,
                    'columns': sorted(list(all_columns)),
                    'null_counts': null_counts,
                    'unique_counts': {col: len(vals) for col, vals in unique_values.items()},
                    'statistics': {}
                }
                
                # Store analysis
                await conn.execute("""
                    INSERT INTO dsa_core.table_analysis (commit_id, table_key, analysis)
                    VALUES ($1, $2, $3)
                    ON CONFLICT (commit_id, table_key) 
                    DO UPDATE SET analysis = $3, created_at = NOW()
                """, commit_id, table_key, json.dumps(analysis))
                
                # Also store schema in commit_schemas for compatibility
                schema_data = {
                    table_key: {
                        "columns": [{"name": col, "type": column_types.get(col, "string")} 
                                   for col in sorted(all_columns)]
                    }
                }
                
                # Check if schema already exists for this commit
                existing_schema = await conn.fetchval("""
                    SELECT schema_definition 
                    FROM dsa_core.commit_schemas 
                    WHERE commit_id = $1
                """, commit_id)
                
                if existing_schema:
                    # Update existing schema
                    if isinstance(existing_schema, str):
                        existing_schema = json.loads(existing_schema)
                    existing_schema.update(schema_data)
                    await conn.execute("""
                        UPDATE dsa_core.commit_schemas 
                        SET schema_definition = $2
                        WHERE commit_id = $1
                    """, commit_id, json.dumps(existing_schema))
                else:
                    # Insert new schema
                    await conn.execute("""
                        INSERT INTO dsa_core.commit_schemas (commit_id, schema_definition)
                        VALUES ($1, $2)
                    """, commit_id, json.dumps(schema_data))
    
    async def _run_post_import_maintenance(self, commit_id: str, job_id: str, db_pool: DatabasePool) -> None:
        """Run database maintenance tasks after a successful import."""
        import logging
        logger = logging.getLogger(__name__)
        
        # Analyze imported tables
        logger.info(f"Import job {job_id} - Analyzing imported tables")
        await self._analyze_imported_tables(commit_id, db_pool)
        
        # Run VACUUM ANALYZE on the affected tables
        logger.info(f"Import job {job_id} - Running VACUUM ANALYZE on core tables")
        async with db_pool.acquire() as conn:
            # Using a longer statement timeout for potentially long-running maintenance
            await conn.execute("SET statement_timeout = '30min';")
            await conn.execute("VACUUM (VERBOSE, ANALYZE) dsa_core.rows;")
            await conn.execute("VACUUM (VERBOSE, ANALYZE) dsa_core.commit_rows;")
        
        # Refresh materialized views
        logger.info(f"Import job {job_id} - Refreshing search index")
        async with db_pool.acquire() as conn:
            await conn.execute("REFRESH MATERIALIZED VIEW CONCURRENTLY dsa_search.datasets_summary;")
    
    async def _find_line_boundaries(self, file_path: str, num_chunks: int, start_offset: int = 0) -> List[Tuple[int, int]]:
        """
        Finds byte positions and line numbers for clean file splitting in a single pass.
        """
        file_size = await aiofiles.os.path.getsize(file_path)
        data_size = file_size - start_offset
        if data_size <= 0:
            return [(start_offset, 2), (file_size, -1)]
        
        # Calculate target byte offsets for each chunk start
        chunk_size = data_size // num_chunks
        target_offsets = {start_offset + (i * chunk_size) for i in range(1, num_chunks)}
        
        boundaries = [(start_offset, 2)]  # First chunk always starts at offset and line 2
        next_target = min(target_offsets) if target_offsets else file_size + 1
        
        current_pos = start_offset
        line_count = 2
        
        async with aiofiles.open(file_path, 'rb') as f:
            await f.seek(start_offset)
            
            # Use a large, efficient read buffer
            read_buffer_size = 16 * 1024 * 1024  # 16MB
            
            while current_pos < file_size and target_offsets:
                chunk = await f.read(read_buffer_size)
                if not chunk:
                    break
                
                chunk_start_pos = current_pos
                
                # While there are targets within our current chunk
                while next_target < chunk_start_pos + len(chunk):
                    # Find the position of the target relative to the start of the chunk
                    relative_pos = int(next_target - chunk_start_pos)
                    
                    # Count newlines up to that relative position
                    lines_in_prefix = chunk[:relative_pos].count(b'\n')
                    
                    # Find the next newline *after* the target position
                    newline_offset = chunk.find(b'\n', relative_pos)
                    
                    if newline_offset != -1:
                        # Found a clean break point
                        absolute_boundary = chunk_start_pos + newline_offset + 1
                        lines_up_to_boundary = line_count + chunk[:newline_offset + 1].count(b'\n')
                        
                        boundaries.append((absolute_boundary, lines_up_to_boundary))
                        
                        # Remove the target we just processed and find the next one
                        target_offsets.discard(next_target)
                        next_target = min(target_offsets) if target_offsets else file_size + 1
                    else:
                        # No newline found in the rest of this chunk, break to read more data
                        break
                
                # Update position and line count for the next iteration
                line_count += chunk.count(b'\n')
                current_pos += len(chunk)

        # De-duplicate and sort, just in case
        final_boundaries = sorted(list(set(boundaries)), key=lambda x: x[0])
        
        # Ensure the last boundary points to the end of the file
        if not final_boundaries or final_boundaries[-1][0] < file_size:
            final_boundaries.append((file_size, -1))  # Line number for end doesn't matter

        return final_boundaries
    
    async def _should_use_parallel_processing(self, file_path: str) -> bool:
        """Determine if parallel processing would be beneficial based on file size."""
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        return file_size_mb > self.parallel_threshold_mb and self.parallel_workers > 1
    
    async def _process_csv_file_parallel(
        self, 
        commit_id: str,
        file_path: str, 
        job_id: str, 
        db_pool: DatabasePool
    ) -> int:
        """Process large CSV files in parallel using Polars for high performance."""
        import logging
        import polars as pl
        
        logger = logging.getLogger(__name__)
        logger.info(f"Import job {job_id} - Starting parallel CSV processing with {self.parallel_workers} workers")

        # 1. Get total row count and column names using Polars (very fast)
        loop = asyncio.get_running_loop()
        
        def get_csv_info():
            # Quick scan to get row count and columns
            df = pl.scan_csv(file_path).select(pl.count()).collect()
            total_rows = df[0, 0]
            # Get column names from the first row
            columns = pl.read_csv(file_path, n_rows=1).columns
            return total_rows, columns
        
        total_rows, column_names = await loop.run_in_executor(None, get_csv_info)
        logger.info(f"Import job {job_id} - Found {total_rows:,} rows to process")
        
        # 2. Calculate row ranges for each worker
        rows_per_worker = total_rows // self.parallel_workers
        boundaries = []
        for i in range(self.parallel_workers):
            start_row = i * rows_per_worker
            if i == self.parallel_workers - 1:
                # Last worker handles any remaining rows
                num_rows = total_rows - start_row
            else:
                num_rows = rows_per_worker
            boundaries.append((start_row, num_rows))
        
        total_rows_processed = 0
        
        # 3. Use a Queue for progress reporting
        mp_manager = mp.Manager()
        progress_queue = mp_manager.Queue()

        # 4. Create a listener task for the queue in the main async process
        async def progress_listener():
            nonlocal total_rows_processed
            while True:
                # A value of -1 signals the end
                processed_count = await asyncio.to_thread(progress_queue.get)
                if processed_count == -1:
                    return
                total_rows_processed += processed_count
                await self._update_job_progress(
                    job_id,
                    {"status": f"Processing: {total_rows_processed:,} rows", "rows_processed": total_rows_processed},
                    db_pool
                )

        listener_task = asyncio.create_task(progress_listener())
        
        try:
            with ProcessPoolExecutor(max_workers=self.parallel_workers) as executor:
                futures = []
                for i, (start_row, num_rows) in enumerate(boundaries):
                    future = executor.submit(
                        _process_csv_chunk_worker_polars,
                        file_path, column_names, start_row, num_rows,
                        commit_id, self.db_url, self.batch_size, 
                        self.use_xxhash, self.xxhash_seed, progress_queue, i
                    )
                    futures.append(future)

                # Wait for all futures to complete
                for future in as_completed(futures):
                    try:
                        # result() will re-raise exceptions from the worker
                        chunk_rows = future.result() 
                    except Exception as e:
                        logger.error(f"Worker failed: {e}")
                        raise
                
                # Signal the listener to stop
                progress_queue.put(-1)
                await listener_task

        except Exception as e:
            listener_task.cancel()  # Ensure listener is cancelled on error
            logger.error(f"A worker failed during parallel processing: {e}")
            raise

        logger.info(f"Import job {job_id} - Parallel processing completed. Total rows: {total_rows_processed:,}")
        return total_rows_processed


# Module-level worker functions (must be at module level for multiprocessing)
def _commit_batch_sync_refactored(conn, batch: List[Tuple[int, Dict]], commit_id: str, hash_func):
    """Refactored commit function using efficient in-memory buffer for COPY."""
    if not batch:
        return
    
    sheet_name = 'primary'
    buffer = io.StringIO()

    for line_num, row_data in batch:
        data_json_bytes = json.dumps(row_data, sort_keys=True, separators=(',', ':')).encode('utf-8')
        data_hash = hash_func(data_json_bytes)
        logical_row_id = f"{sheet_name}:{line_num}"  # Use line number instead of hash
        
        # Store only the raw row data, metadata is in logical_row_id
        data_json = data_json_bytes.decode('utf-8')
        
        # Write tab-separated values to the buffer
        buffer.write(f"{logical_row_id}\t{data_hash}\t{data_json}\n")

    buffer.seek(0)
    
    with conn.cursor() as cur:
        # Clear the temp table for the new batch
        cur.execute("TRUNCATE import_batch;")
        
        with cur.copy("COPY import_batch (logical_row_id, row_hash, data) FROM STDIN") as copy:
            copy.write(buffer.read())
        
        with conn.transaction():
            cur.execute("INSERT INTO dsa_core.rows (row_hash, data) SELECT row_hash, data FROM import_batch ON CONFLICT (row_hash) DO NOTHING;")
            cur.execute("INSERT INTO dsa_core.commit_rows (commit_id, logical_row_id, row_hash) SELECT %s, logical_row_id, row_hash FROM import_batch;", (commit_id,))


# New Polars-based worker function
def _process_csv_chunk_worker_polars(
    file_path: str, column_names: List[str], start_row: int, num_rows: int,
    commit_id: str, db_url: str, batch_size: int,
    use_xxhash: bool, xxhash_seed: int, progress_queue: mp.Queue, worker_id: int
) -> int:
    """Process CSV chunk using Polars for maximum performance."""
    import polars as pl
    import json, xxhash, hashlib, psycopg, logging, io
    
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(f"worker_{worker_id}")
    
    def calculate_hash(data: bytes) -> str:
        if use_xxhash:
            return xxhash.xxh64(data, seed=xxhash_seed).hexdigest()
        else:
            return hashlib.sha256(data).hexdigest()
    
    try:
        # Use Polars to read specific rows from CSV - extremely fast
        df_chunk = pl.read_csv(
            file_path,
            skip_rows=start_row + 1,  # +1 to skip header
            n_rows=num_rows,
            has_header=False,
            new_columns=column_names  # Apply the column names
        )
        
        total_rows = 0
        batch_with_line_nums = []
        # Global line numbers start at 2 (after header) + start_row
        current_line = start_row + 2
        
        with psycopg.connect(db_url) as conn:
            # Create the temp table ONCE per worker
            with conn.cursor() as cur:
                # Optimize session for bulk loading
                cur.execute("SET work_mem = '256MB';")
                cur.execute("SET maintenance_work_mem = '256MB';")
                cur.execute("SET synchronous_commit = OFF;")
                # TEMP tables are already unlogged and use temp_buffers
                cur.execute("CREATE TEMP TABLE import_batch (logical_row_id TEXT, row_hash TEXT, data JSONB) ON COMMIT DROP;")
            
            # Use iter_rows for memory efficiency
            for row_data in df_chunk.iter_rows(named=True):
                batch_with_line_nums.append((current_line, row_data))
                current_line += 1
                
                if len(batch_with_line_nums) >= batch_size:
                    _commit_batch_sync_refactored(conn, batch_with_line_nums, commit_id, calculate_hash)
                    progress_queue.put(len(batch_with_line_nums))
                    total_rows += len(batch_with_line_nums)
                    batch_with_line_nums = []
            
            if batch_with_line_nums:
                _commit_batch_sync_refactored(conn, batch_with_line_nums, commit_id, calculate_hash)
                progress_queue.put(len(batch_with_line_nums))
                total_rows += len(batch_with_line_nums)
        
        logger.info(f"Worker {worker_id} completed successfully. Processed {total_rows} rows.")
        return total_rows
        
    except Exception as e:
        logger.error(f"Worker {worker_id} failed: {e}", exc_info=True)
        raise