import duckdb
import logging
import asyncio
import tempfile
import os
from io import BytesIO
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from app.sampling.models import (
    SamplingMethod, JobStatus, SamplingRequest, 
    SamplingJob, RandomSamplingParams, StratifiedSamplingParams,
    SystematicSamplingParams, ClusterSamplingParams, CustomSamplingParams,
    DataFilters, DataSelection, FilterCondition, DataSummary
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class SamplingService:
    def __init__(self, datasets_repository, sampling_repository):
        self.datasets_repository = datasets_repository
        self.sampling_repository = sampling_repository
        
    async def create_sampling_job(
        self, 
        dataset_id: int,
        version_id: int,
        request: SamplingRequest,
        user_id: int
    ) -> SamplingJob:
        """
        Create and enqueue a new sampling job
        
        Args:
            dataset_id: ID of the dataset to sample
            version_id: Version of the dataset to sample
            request: Sampling request with method and parameters
            user_id: ID of the user creating the job
            
        Returns:
            A SamplingJob object with a unique ID
        """
        # Create a new job
        job = SamplingJob(
            dataset_id=dataset_id,
            version_id=version_id,
            user_id=user_id,
            request=request
        )
        
        # Store the job
        await self.sampling_repository.create_job(job)
        
        # Start the job in the background
        asyncio.create_task(self._process_job(job.id))
        
        return job
    
    async def get_job(self, job_id: str) -> Optional[SamplingJob]:
        """Get job details by ID"""
        return await self.sampling_repository.get_job(job_id)
    
    async def get_job_preview(self, job_id: str) -> List[Dict[str, Any]]:
        """Get preview data for a job"""
        job = await self.sampling_repository.get_job(job_id)
        if not job:
            return []
        
        return job.output_preview or []
    
    async def get_dataset_columns(self, dataset_id: int, version_id: int) -> Dict[str, Any]:
        """Get column information for a dataset version"""
        try:
            # Create a new connection for this operation
            from app.db.connection import AsyncSessionLocal
            
            async with AsyncSessionLocal() as session:
                from app.datasets.repository import DatasetsRepository
                datasets_repo = DatasetsRepository(session)
                
                # Get dataset version
                version = await datasets_repo.get_dataset_version(version_id)
                if not version:
                    raise ValueError(f"Dataset version with ID {version_id} not found")
                
                # Verify dataset ID matches
                if version.dataset_id != dataset_id:
                    raise ValueError(f"Version {version_id} does not belong to dataset {dataset_id}")
                
                # Get file data
                file_info = await datasets_repo.get_file(version.file_id)
                if not file_info or not file_info.file_data:
                    raise ValueError("File data not found")
                
                # Create DuckDB connection and load data
                conn = duckdb.connect(':memory:')
                temp_file_path = self._load_data_to_duckdb(conn, file_info)
                
                try:
                    # Get column information
                    data_summary = self._get_data_summary(conn, 'main_data')
                    
                    # Get sample of unique values for each column (helpful for filters)
                    sample_values = {}
                    for col_name in data_summary.column_types.keys():
                        try:
                            # Get first 10 unique values for this column
                            result = conn.execute(f'''
                                SELECT DISTINCT "{col_name}" 
                                FROM main_data 
                                WHERE "{col_name}" IS NOT NULL 
                                ORDER BY "{col_name}" 
                                LIMIT 10
                            ''').fetchall()
                            sample_values[col_name] = [row[0] for row in result]
                        except Exception:
                            # Skip columns that can't be sampled (e.g., complex types)
                            sample_values[col_name] = []
                    
                    return {
                        "columns": list(data_summary.column_types.keys()),
                        "column_types": data_summary.column_types,
                        "total_rows": data_summary.total_rows,
                        "null_counts": data_summary.null_counts,
                        "sample_values": sample_values
                    }
                finally:
                    # Clean up temp file
                    if temp_file_path and os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                
        except Exception as e:
            logger.error(f"Error getting dataset columns: {str(e)}", exc_info=True)
            raise ValueError(f"Error getting dataset columns: {str(e)}")
    
    async def _process_job(self, job_id: str) -> None:
        """
        Process a sampling job in the background
        
        This method loads the dataset, applies the sampling method,
        and updates the job status.
        """
        job = await self.sampling_repository.get_job(job_id)
        if not job:
            logger.error(f"Job {job_id} not found")
            return
        
        try:
            # Update job status
            job.status = JobStatus.RUNNING
            job.started_at = datetime.now()
            await self.sampling_repository.update_job(job)
            
            # Validate dataset and version
            # Create a new connection for this background task
            from app.db.connection import AsyncSessionLocal
            
            async with AsyncSessionLocal() as session:
                from app.datasets.repository import DatasetsRepository
                datasets_repo = DatasetsRepository(session)
                
                # Get dataset version
                version = await datasets_repo.get_dataset_version(job.version_id)
                if not version:
                    raise ValueError(f"Dataset version with ID {job.version_id} not found")
                
                # Verify dataset ID matches
                if version.dataset_id != job.dataset_id:
                    raise ValueError(f"Version {job.version_id} does not belong to dataset {job.dataset_id}")
                
                # Get file data
                file_info = await datasets_repo.get_file(version.file_id)
                if not file_info or not file_info.file_data:
                    raise ValueError("File data not found")
                
                # Create DuckDB connection
                conn = duckdb.connect(':memory:')
                temp_file_path = None
                
                try:
                    # Load file into DuckDB
                    temp_file_path = self._load_data_to_duckdb(conn, file_info, job.request.sheet)
                    
                    # Generate data summary
                    data_summary = self._get_data_summary(conn, 'main_data')
                    
                    # Apply filtering and sampling
                    sampled_data = await self._apply_sampling_with_duckdb(conn, job.request)
                    
                    # Generate sample summary
                    sample_summary = self._get_sample_summary_from_duckdb(conn, sampled_data)
                    
                    # Get preview data
                    preview_result = conn.execute(f"SELECT * FROM ({sampled_data}) LIMIT 10").fetchall()
                    columns = [desc[0] for desc in conn.description]
                    job.output_preview = [dict(zip(columns, row)) for row in preview_result]
                    
                    job.data_summary = data_summary
                    job.sample_summary = sample_summary
                    
                    # Use a local file path for the mock URI
                    job.output_uri = f"file://outputs/samples/{job.dataset_id}/{job.version_id}/{job_id}.parquet"

                    # Update job status
                    job.status = JobStatus.COMPLETED
                    job.completed_at = datetime.now()
                    await self.sampling_repository.update_job(job)
                finally:
                    # Clean up temp file
                    if temp_file_path and os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
            
        except Exception as e:
            # Handle job failure
            logger.error(f"Error processing job {job_id}: {str(e)}", exc_info=True)
            job.status = JobStatus.FAILED
            job.error_message = str(e)
            if not job.started_at:
                job.started_at = datetime.now()
            job.completed_at = datetime.now()
            await self.sampling_repository.update_job(job)
    
    def _create_temp_file_from_bytes(self, file_data: bytes, file_type: str) -> str:
        """Create a temporary file from bytes data and return the path"""
        suffix = f".{file_type.lower()}"
        if file_type.lower() in ["xls", "xlsx", "xlsm"]:
            suffix = ".xlsx"  # Excel files
        elif file_type.lower() == "parquet":
            suffix = ".parquet"
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        try:
            temp_file.write(file_data)
            temp_file.flush()
            temp_file.close()
            return temp_file.name
        except Exception as e:
            # Clean up on error
            temp_file.close()
            os.unlink(temp_file.name)
            raise e
    
    async def _apply_sampling_sql(self, conn: duckdb.DuckDBPyConnection, base_query: str, request: SamplingRequest) -> str:
        """
        Generate SQL query for the requested sampling method
        
        Args:
            conn: DuckDB connection
            base_query: Base SQL query to sample from
            request: Sampling request with method and parameters
            
        Returns:
            SQL query string for the sampled data
        """
        try:
            # Get typed parameters
            params = request.get_typed_parameters()
            
            # Create temporary view from base query
            conn.execute(f"CREATE OR REPLACE TEMPORARY VIEW filtered_data AS {base_query}")
            
            # Apply sampling method
            if request.method == SamplingMethod.RANDOM:
                return self._random_sampling_sql(conn, params)
            elif request.method == SamplingMethod.STRATIFIED:
                return self._stratified_sampling_sql(conn, params)
            elif request.method == SamplingMethod.SYSTEMATIC:
                return self._systematic_sampling_sql(conn, params)
            elif request.method == SamplingMethod.CLUSTER:
                return self._cluster_sampling_sql(conn, params)
            elif request.method == SamplingMethod.CUSTOM:
                return self._custom_sampling_sql(conn, params)
            else:
                raise ValueError(f"Unknown sampling method: {request.method}")
        except Exception as e:
            logger.error(f"Error applying sampling: {str(e)}", exc_info=True)
            raise ValueError(f"Error applying sampling: {str(e)}")
    
    def _random_sampling_sql(self, conn: duckdb.DuckDBPyConnection, params: RandomSamplingParams) -> str:
        """Generate SQL for random sampling"""
        # Get total count
        total_count = conn.execute("SELECT COUNT(*) FROM filtered_data").fetchone()[0]
        
        if params.sample_size >= total_count:
            return "SELECT * FROM filtered_data"
        
        # Use DuckDB's SAMPLE clause with seed if provided
        if params.seed is not None:
            return f"SELECT * FROM filtered_data USING SAMPLE {params.sample_size} ROWS (SYSTEM, {params.seed})"
        else:
            return f"SELECT * FROM filtered_data USING SAMPLE {params.sample_size} ROWS"
    
    def _stratified_sampling_sql(self, conn: duckdb.DuckDBPyConnection, params: StratifiedSamplingParams) -> str:
        """Generate SQL for stratified sampling"""
        # Validate strata columns exist
        columns_result = conn.execute("PRAGMA table_info('filtered_data')").fetchall()
        available_columns = [col[1] for col in columns_result]
        
        for col in params.strata_columns:
            if col not in available_columns:
                raise ValueError(f"Strata column '{col}' not found in dataset")
        
        # Build strata expression
        strata_expr = " || '_' || ".join([f'CAST("{col}" AS VARCHAR)' for col in params.strata_columns])
        
        # Get stratum counts
        strata_query = f"""
        SELECT {strata_expr} as stratum, COUNT(*) as cnt 
        FROM filtered_data 
        GROUP BY stratum
        """
        strata_counts = conn.execute(strata_query).fetchall()
        total_rows = sum(count for _, count in strata_counts)
        
        # Build sampling queries per stratum
        if params.sample_size is None and params.min_per_stratum is None:
            # Default to 10% per stratum
            fraction = 0.1
            sample_queries = []
            for stratum, count in strata_counts:
                n_samples = max(1, int(count * fraction))
                if params.seed:
                    sample_queries.append(f"""
                    SELECT * FROM filtered_data 
                    WHERE {strata_expr} = '{stratum}' 
                    USING SAMPLE {n_samples} ROWS (SYSTEM, {params.seed})
                    """)
                else:
                    sample_queries.append(f"""
                    SELECT * FROM filtered_data 
                    WHERE {strata_expr} = '{stratum}' 
                    USING SAMPLE {n_samples} ROWS
                    """)
        elif isinstance(params.sample_size, float):
            # Sample by fraction
            fraction = params.sample_size
            sample_queries = []
            for stratum, count in strata_counts:
                n_samples = max(1, int(count * fraction))
                if params.seed:
                    sample_queries.append(f"""
                    SELECT * FROM filtered_data 
                    WHERE {strata_expr} = '{stratum}' 
                    USING SAMPLE {n_samples} ROWS (SYSTEM, {params.seed})
                    """)
                else:
                    sample_queries.append(f"""
                    SELECT * FROM filtered_data 
                    WHERE {strata_expr} = '{stratum}' 
                    USING SAMPLE {n_samples} ROWS
                    """)
        else:
            # Proportional allocation with minimum
            total_samples = params.sample_size if params.sample_size else int(total_rows * 0.1)
            sample_queries = []
            
            for stratum, count in strata_counts:
                allocated = max(
                    int(total_samples * (count / total_rows)),
                    params.min_per_stratum or 0
                )
                n_samples = min(allocated, count)
                
                if n_samples > 0:
                    if params.seed:
                        sample_queries.append(f"""
                        SELECT * FROM filtered_data 
                        WHERE {strata_expr} = '{stratum}' 
                        USING SAMPLE {n_samples} ROWS (SYSTEM, {params.seed})
                        """)
                    else:
                        sample_queries.append(f"""
                        SELECT * FROM filtered_data 
                        WHERE {strata_expr} = '{stratum}' 
                        USING SAMPLE {n_samples} ROWS
                        """)
        
        # Combine all queries with UNION ALL
        return " UNION ALL ".join(f"({q})" for q in sample_queries)
    
    def _systematic_sampling_sql(self, conn: duckdb.DuckDBPyConnection, params: SystematicSamplingParams) -> str:
        """Generate SQL for systematic sampling"""
        if params.interval <= 0:
            raise ValueError("Interval must be greater than 0")
        
        start = params.start if params.start is not None else 0
        
        # Use ROW_NUMBER() to implement systematic sampling
        return f"""
        SELECT * FROM (
            SELECT *, ROW_NUMBER() OVER () - 1 as rn 
            FROM filtered_data
        ) t 
        WHERE (rn - {start}) % {params.interval} = 0
        """
    
    def _cluster_sampling_sql(self, conn: duckdb.DuckDBPyConnection, params: ClusterSamplingParams) -> str:
        """Generate SQL for cluster sampling"""
        # Validate cluster column exists
        columns_result = conn.execute("PRAGMA table_info('filtered_data')").fetchall()
        available_columns = [col[1] for col in columns_result]
        
        if params.cluster_column not in available_columns:
            raise ValueError(f"Cluster column '{params.cluster_column}' not found in dataset")
        
        # Get unique clusters
        clusters_result = conn.execute(f'SELECT DISTINCT "{params.cluster_column}" FROM filtered_data').fetchall()
        clusters = [row[0] for row in clusters_result]
        
        if params.num_clusters >= len(clusters):
            # If we want more clusters than exist, return all
            return "SELECT * FROM filtered_data"
        
        # Sample clusters randomly
        import random
        sampled_clusters = random.sample(clusters, params.num_clusters)
        
        # Build IN clause with proper escaping
        cluster_values = ", ".join([f"'{c}'" if isinstance(c, str) else str(c) for c in sampled_clusters])
        
        base_query = f'SELECT * FROM filtered_data WHERE "{params.cluster_column}" IN ({cluster_values})'
        
        # Optionally sample within clusters
        if params.sample_within_clusters:
            # Sample 50% within each cluster using window functions
            return f"""
            WITH ranked AS (
                SELECT *, 
                       ROW_NUMBER() OVER (PARTITION BY "{params.cluster_column}" ORDER BY RANDOM()) as rn,
                       COUNT(*) OVER (PARTITION BY "{params.cluster_column}") as cluster_size
                FROM ({base_query}) t
            )
            SELECT * FROM ranked WHERE rn <= cluster_size / 2
            """
        
        return base_query
    
    def _custom_sampling_sql(self, conn: duckdb.DuckDBPyConnection, params: CustomSamplingParams) -> str:
        """Generate SQL for custom sampling"""
        # The custom query parameter should contain a WHERE clause condition
        # We'll wrap it in a proper SQL query
        return f"SELECT * FROM filtered_data WHERE {params.query}"
    
    def _load_data_to_duckdb(self, conn: duckdb.DuckDBPyConnection, file_info: Any, sheet_name: Optional[str] = None) -> str:
        """Load file data into DuckDB table and return temp file path (if created)"""
        file_data = file_info.file_data
        file_type = file_info.file_type.lower()
        temp_file_path = None
        
        try:
            # Create a temporary file from the bytes data
            temp_file_path = self._create_temp_file_from_bytes(file_data, file_type)
            
            if file_type == "csv":
                # Use DuckDB's read_csv_auto for automatic CSV parsing
                conn.execute(f"CREATE TABLE main_data AS SELECT * FROM read_csv_auto('{temp_file_path}')")
            elif file_type == "parquet":
                # Use DuckDB's read_parquet for Parquet files
                conn.execute(f"CREATE TABLE main_data AS SELECT * FROM read_parquet('{temp_file_path}')")
            elif file_type in ["xls", "xlsx", "xlsm"]:
                # For Excel files, we need to convert to CSV first
                # DuckDB doesn't have native Excel support
                import openpyxl
                import csv
                
                # Load the workbook
                wb = openpyxl.load_workbook(temp_file_path, read_only=True)
                
                # Get the sheet
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active  # Use the first sheet
                
                # Create a temporary CSV file
                csv_temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', newline='')
                csv_writer = csv.writer(csv_temp_file)
                
                try:
                    # Write all rows to CSV
                    for row in ws.iter_rows(values_only=True):
                        csv_writer.writerow(row)
                    csv_temp_file.close()
                    
                    # Now load the CSV into DuckDB
                    conn.execute(f"CREATE TABLE main_data AS SELECT * FROM read_csv_auto('{csv_temp_file.name}')")
                    
                    # Clean up the CSV file
                    os.unlink(csv_temp_file.name)
                finally:
                    wb.close()
            else:
                # Fall back to CSV for unknown types
                conn.execute(f"CREATE TABLE main_data AS SELECT * FROM read_csv_auto('{temp_file_path}')")
            
            return temp_file_path
            
        except Exception as e:
            # Clean up temp file on error
            if temp_file_path and os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
            logger.error(f"Error loading file to DuckDB: {str(e)}")
            raise ValueError(f"Error loading file: {str(e)}")
    
    def _get_data_summary(self, conn: duckdb.DuckDBPyConnection, table_name: str) -> DataSummary:
        """Generate data summary statistics using DuckDB"""
        try:
            # Get basic info
            result = conn.execute(f"SELECT COUNT(*) as total_rows FROM {table_name}").fetchone()
            total_rows = result[0]
            
            # Get column info
            columns_result = conn.execute(f"PRAGMA table_info('{table_name}')").fetchall()
            total_columns = len(columns_result)
            
            # Get column types and null counts
            column_types = {}
            null_counts = {}
            
            for col_info in columns_result:
                col_name = col_info[1]
                col_type = col_info[2]
                column_types[col_name] = col_type
                
                # Get null count for this column
                null_result = conn.execute(f"SELECT COUNT(*) FROM {table_name} WHERE \"{col_name}\" IS NULL").fetchone()
                null_counts[col_name] = null_result[0]
            
            # Estimate memory usage (rough calculation)
            memory_usage_mb = total_rows * total_columns * 8 / (1024 * 1024)  # Rough estimate
            
            return DataSummary(
                total_rows=total_rows,
                total_columns=total_columns,
                column_types=column_types,
                memory_usage_mb=round(memory_usage_mb, 2),
                null_counts=null_counts
            )
            
        except Exception as e:
            logger.error(f"Error getting data summary: {str(e)}")
            raise ValueError(f"Error getting data summary: {str(e)}")
    
    async def execute_sampling_synchronously(
        self,
        dataset_id: int,
        version_id: int,
        request: SamplingRequest
    ) -> List[Dict[str, Any]]:
        """
        Execute sampling synchronously and return the result as a list of dictionaries.
        """
        try:
            # Create a new connection for this operation
            from app.db.connection import AsyncSessionLocal

            async with AsyncSessionLocal() as session:
                from app.datasets.repository import DatasetsRepository
                datasets_repo = DatasetsRepository(session)

                # Get dataset version
                version = await datasets_repo.get_dataset_version(version_id)
                if not version:
                    raise ValueError(f"Dataset version with ID {version_id} not found")

                # Verify dataset ID matches
                if version.dataset_id != dataset_id:
                    raise ValueError(f"Version {version_id} does not belong to dataset {dataset_id}")

                # Get file data
                file_info = await datasets_repo.get_file(version.file_id)
                if not file_info or not file_info.file_data:
                    raise ValueError("File data not found")

                # Create DuckDB connection
                conn = duckdb.connect(':memory:')
                temp_file_path = None
                
                try:
                    # Load file into DuckDB
                    temp_file_path = self._load_data_to_duckdb(conn, file_info, request.sheet)

                    # Apply filtering and sampling
                    sampled_query = await self._apply_sampling_with_duckdb(conn, request)
                    
                    # Create a temporary table with the sampled data
                    conn.execute(f"CREATE OR REPLACE TEMPORARY TABLE sampled_result AS {sampled_query}")
                    
                    # Fetch all data from the sampled result
                    result = conn.execute("SELECT * FROM sampled_result").fetchall()
                    columns = [desc[0] for desc in conn.description]
                    
                    # Return as list of dictionaries (similar to DataFrame.to_dict('records'))
                    return [dict(zip(columns, row)) for row in result]
                    
                finally:
                    # Clean up temp file
                    if temp_file_path and os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)

        except Exception as e:
            logger.error(f"Error executing sampling synchronously: {str(e)}", exc_info=True)
            # Re-raise as ValueError to be handled by the controller
            raise ValueError(f"Error executing sampling synchronously: {str(e)}")

    def _get_sample_summary_from_duckdb(self, conn: duckdb.DuckDBPyConnection, sample_query: str) -> DataSummary:
        """Generate summary statistics for the sampled data using DuckDB"""
        try:
            # Create temporary view for the sample
            conn.execute(f"CREATE OR REPLACE TEMPORARY VIEW sample_data AS {sample_query}")
            
            # Get basic info
            result = conn.execute("SELECT COUNT(*) as total_rows FROM sample_data").fetchone()
            total_rows = result[0]
            
            # Get column info
            columns_result = conn.execute("PRAGMA table_info('sample_data')").fetchall()
            total_columns = len(columns_result)
            
            # Get column types and null counts
            column_types = {}
            null_counts = {}
            
            for col_info in columns_result:
                col_name = col_info[1]
                col_type = col_info[2]
                column_types[col_name] = col_type
                
                # Get null count for this column
                null_result = conn.execute(f'SELECT COUNT(*) FROM sample_data WHERE "{col_name}" IS NULL').fetchone()
                null_counts[col_name] = null_result[0]
            
            # Estimate memory usage (rough calculation)
            memory_usage_mb = total_rows * total_columns * 8 / (1024 * 1024)  # Rough estimate
            
            return DataSummary(
                total_rows=total_rows,
                total_columns=total_columns,
                column_types=column_types,
                memory_usage_mb=round(memory_usage_mb, 2),
                null_counts=null_counts
            )
            
        except Exception as e:
            logger.error(f"Error getting sample summary: {str(e)}")
            raise ValueError(f"Error getting sample summary: {str(e)}")
    
    def _build_filter_query(self, filters: Optional[DataFilters]) -> Tuple[str, List[Any]]:
        """Build SQL WHERE clause from filter conditions, returning clause and parameters."""
        if not filters or not filters.conditions:
            return "", []

        condition_strings = []
        params: List[Any] = []
        for condition in filters.conditions:
            col = f'"{condition.column}"'  # Quote column names

            if condition.operator in ['IS NULL', 'IS NOT NULL']:
                condition_strings.append(f"{col} {condition.operator}")
            elif condition.operator in ['IN', 'NOT IN']:
                if isinstance(condition.value, list):
                    if not condition.value:  # Empty list
                        if condition.operator == 'IN':
                            condition_strings.append("0=1")  # Always false for IN empty list
                        else:  # NOT IN
                            condition_strings.append("1=1")  # Always true for NOT IN empty list
                    else:  # Non-empty list
                        placeholders = ', '.join(['?'] * len(condition.value))
                        condition_strings.append(f"{col} {condition.operator} ({placeholders})")
                        params.extend(condition.value)
                else:  # Single value, treat as = or !=
                    actual_operator = '=' if condition.operator == 'IN' else '!='
                    condition_strings.append(f"{col} {actual_operator} ?")
                    params.append(condition.value)
            else:  # For other operators like =, !=, >, <, LIKE, ILIKE
                condition_strings.append(f"{col} {condition.operator} ?")
                params.append(condition.value)

        if not condition_strings:
            return "", []

        return f"WHERE {f' {filters.logic} '.join(condition_strings)}", params

    def _build_select_from_clause(self, conn: duckdb.DuckDBPyConnection, selection: Optional[DataSelection], table_name: str = 'main_data') -> str:
        """Build SQL SELECT ... FROM ... clause, handling column selection and exclusion."""
        columns_sql = "*"
        if selection:
            if selection.columns:
                columns_sql = ', '.join([f'"{col}"' for col in selection.columns])
            elif selection.exclude_columns:
                # Fetch all column names from the table
                all_columns_info = conn.execute(f"PRAGMA table_info('{table_name}')").fetchall()
                all_column_names = [info[1] for info in all_columns_info]

                # Filter out excluded columns
                selected_columns = [col for col in all_column_names if col not in selection.exclude_columns]

                if not selected_columns:
                    # If all columns are excluded, this would result in an error.
                    # Default to selecting a literal to prevent SQL errors, or raise.
                    # For simplicity, let's raise if it results in no columns.
                    # A more sophisticated approach might select 'NULL' or similar if allowed.
                    raise ValueError("Excluding all columns or resulting in an empty column set is not allowed.")
                columns_sql = ', '.join([f'"{col}"' for col in selected_columns])

        return f"SELECT {columns_sql} FROM {table_name}"

    def _build_order_limit_offset_clause(self, selection: Optional[DataSelection]) -> str:
        """Build SQL ORDER BY, LIMIT, OFFSET clause."""
        parts = []
        if selection:
            if selection.order_by:
                order_direction = "DESC" if selection.order_desc else "ASC"
                parts.append(f'ORDER BY "{selection.order_by}" {order_direction}')
            if selection.limit is not None: # Allow 0 as a valid limit
                parts.append(f"LIMIT {selection.limit}")
                if selection.offset is not None: # Offset typically used with limit
                    parts.append(f"OFFSET {selection.offset}")
        return " ".join(parts)

    async def _apply_sampling_with_duckdb(self, conn: duckdb.DuckDBPyConnection, request: SamplingRequest) -> str:
        """Apply filtering, selection, and sampling using DuckDB and return the final SQL query."""
        try:
            # Validate filters and selection (checks if columns exist, etc.)
            if request.filters:
                self._validate_filters(conn, request.filters)
            if request.selection:
                self._validate_selection(conn, request.selection)
            
            # Build query parts
            select_from_clause = self._build_select_from_clause(conn, request.selection, 'main_data')
            filter_clause_str, filter_params = self._build_filter_query(request.filters)
            order_limit_offset_clause_str = self._build_order_limit_offset_clause(request.selection)

            query_parts = [select_from_clause]
            if filter_clause_str:
                query_parts.append(filter_clause_str)
            if order_limit_offset_clause_str:
                query_parts.append(order_limit_offset_clause_str)

            base_query = " ".join(query_parts).strip()

            logger.debug(f"Base DuckDB query: {base_query} with params: {filter_params}")

            # Execute the filter params if any
            if filter_params:
                # Create a prepared statement for the base query
                conn.execute(f"CREATE OR REPLACE TEMPORARY VIEW base_filtered AS {base_query}", filter_params)
                base_query = "SELECT * FROM base_filtered"

            # Now apply sampling method to get the final query
            return await self._apply_sampling_sql(conn, base_query, request)
            
        except Exception as e:
            logger.error(f"Error applying sampling with DuckDB: {str(e)}", exc_info=True)
            raise ValueError(f"Error applying sampling with DuckDB: {str(e)}")

    def _validate_filters(self, conn: duckdb.DuckDBPyConnection, filters: DataFilters) -> None:
        """Validate that filter columns exist and have appropriate types"""
        if not filters.conditions:
            return
        
        # Get available columns
        columns_result = conn.execute("PRAGMA table_info('main_data')").fetchall()
        available_columns = {col[1]: col[2] for col in columns_result}  # name: type
        
        for condition in filters.conditions:
            # Check if column exists
            if condition.column not in available_columns:
                raise ValueError(f"Filter column '{condition.column}' does not exist")
            
            # Basic type validation for certain operators
            col_type = available_columns[condition.column].lower()
            if condition.operator in ['>', '<', '>=', '<='] and 'text' in col_type:
                logger.warning(f"Using numeric comparison operator on text column '{condition.column}'")
    
    def _validate_selection(self, conn: duckdb.DuckDBPyConnection, selection: DataSelection) -> None:
        """Validate that selection columns exist"""
        # Get available columns
        columns_result = conn.execute("PRAGMA table_info('main_data')").fetchall()
        available_columns = [col[1] for col in columns_result]
        
        # Validate columns to include
        if selection.columns:
            for col in selection.columns:
                if col not in available_columns:
                    raise ValueError(f"Selection column '{col}' does not exist")
        
        # Validate columns to exclude
        if selection.exclude_columns:
            for col in selection.exclude_columns:
                if col not in available_columns:
                    raise ValueError(f"Exclude column '{col}' does not exist")
        
        # Validate order by column
        if selection.order_by and selection.order_by not in available_columns:
            raise ValueError(f"Order by column '{selection.order_by}' does not exist")
    
    def get_export_formats(self) -> List[str]:
        """Get supported export formats"""
        # Return hardcoded list for now since config is not available
        return ["csv", "parquet", "json"]
    
    async def export_sample(self, job_id: str, format: str) -> bytes:
        """Export sample data in the specified format"""
        # This is a placeholder for export functionality
        # In a real implementation, you would:
        # 1. Get the job and its sampled data
        # 2. Convert to the requested format
        # 3. Return the bytes
        
        job = await self.sampling_repository.get_job(job_id)
        if not job or job.status != JobStatus.COMPLETED:
            raise ValueError("Job not found or not completed")
        
        if format not in self.get_export_formats():
            raise ValueError(f"Unsupported export format: {format}")
        
        # For now, just return a placeholder
        return b"Export functionality not yet implemented"
