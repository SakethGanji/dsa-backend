from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import os
import logging
import pandas as pd
from io import BytesIO
from fastapi import UploadFile, HTTPException
from app.datasets.repository import DatasetsRepository
from app.datasets.models import (
    Dataset, DatasetCreate, DatasetUpdate, DatasetUploadRequest, DatasetUploadResponse,
    DatasetVersion, DatasetVersionCreate, File, FileCreate,
    Sheet, SheetCreate, SheetInfo, SheetMetadata, Tag
)

logger = logging.getLogger(__name__)

class DatasetsService:
    def __init__(self, repository: DatasetsRepository):
        self.repository = repository

    async def upload_dataset(
        self, 
        file: UploadFile, 
        request: DatasetUploadRequest, 
        user_id: int
    ) -> DatasetUploadResponse:
        """
        Process dataset upload following the flow:
        1. Create/Upsert dataset
        2. Save file
        3. Get next version number
        4. Create dataset version
        5. Update dataset timestamp
        6. Process tags if any
        7. Parse file into sheets
        8. Create sheets and metadata
        9. Return response with dataset info
        """
        # Step 0: Upsert dataset if dataset_id provided, otherwise create new dataset
        dataset_create = DatasetCreate(
            name=request.name,
            description=request.description,
            created_by=user_id,
            tags=request.tags
        )
        
        dataset_id = await self.repository.upsert_dataset(request.dataset_id, dataset_create)
        
        # Step 1: Save file
        contents = await file.read()
        file_size = len(contents)
        file_type = os.path.splitext(file.filename)[1].lower()[1:]  # Remove the dot
        
        file_create = FileCreate(
            storage_type="database",
            file_type=file_type,
            mime_type=file.content_type,
            file_data=contents,
            file_size=file_size
        )
        
        file_id = await self.repository.create_file(file_create)
        
        # Step 2: Get next version number
        version_number = await self.repository.get_next_version_number(dataset_id)
        
        # Step 3: Create dataset version
        version_create = DatasetVersionCreate(
            dataset_id=dataset_id,
            version_number=version_number,
            file_id=file_id,
            uploaded_by=user_id
        )
        
        version_id = await self.repository.create_dataset_version(version_create)
        
        # Step 4: Update dataset timestamp
        await self.repository.update_dataset_timestamp(dataset_id)
        
        # Step 5: Process tags if any
        if request.tags:
            for tag_name in request.tags:
                tag_id = await self.repository.upsert_tag(tag_name)
                await self.repository.create_dataset_tag(dataset_id, tag_id)
        
        # Step 6: Parse file into sheets
        sheet_infos = await self._parse_file_into_sheets(
            file=file, 
            contents=contents, 
            file_type=file_type, 
            version_id=version_id
        )
        
        # Step 8: Return response with dataset info
        return DatasetUploadResponse(
            dataset_id=dataset_id,
            version_id=version_id,
            sheets=sheet_infos
        )

    async def _parse_file_into_sheets(
        self,
        file: UploadFile,
        contents: bytes,
        file_type: str,
        version_id: int
    ) -> List[SheetInfo]:
        """Parse file contents into sheets based on file type"""
        sheet_infos = []
        
        try:
            filename = os.path.basename(file.filename)
            parser = self._get_file_parser(file_type)
            sheet_infos = await parser(contents, filename, version_id)
        except Exception as e:
            # Log the error and create an error sheet
            error_msg = f"Error parsing file: {str(e)}"
            logger.error(error_msg, exc_info=True)
            sheet_infos = await self._create_error_sheet(file, version_id, file_type, error_msg)
            
        return sheet_infos
    
    def _get_file_parser(self, file_type: str):
        """Factory method to get the appropriate file parser"""
        parsers = {
            "xlsx": self._parse_excel_file,
            "xls": self._parse_excel_file,
            "xlsm": self._parse_excel_file,
            "csv": self._parse_csv_file
        }
        return parsers.get(file_type, self._handle_unsupported_file_type)
    
    async def _parse_excel_file(self, contents: bytes, filename: str, version_id: int) -> List[SheetInfo]:
        """Parse Excel file into sheets"""
        sheet_infos = []
        
        # Create a BytesIO object from contents for pandas to read
        excel_buffer = BytesIO(contents)
        excel_file = pd.ExcelFile(excel_buffer)
        
        for i, sheet_name in enumerate(excel_file.sheet_names):
            sheet_create = SheetCreate(
                dataset_version_id=version_id,
                name=sheet_name,
                sheet_index=i,
                description=None
            )

            sheet_id = await self.repository.create_sheet(sheet_create)

            # Optional: Process sheet for profiling
            # profile_report_file_id = await self._generate_profile_report(excel_file, sheet_name)

            # Create sheet metadata
            metadata = {
                "columns": len(excel_file.parse(sheet_name).columns),
                "rows": len(excel_file.parse(sheet_name))
                # Additional metadata can be added here
            }

            await self.repository.create_sheet_metadata(sheet_id, metadata)

            sheet_infos.append(SheetInfo(
                id=sheet_id,
                name=sheet_name,
                index=i,
                description=None
            ))
            
        return sheet_infos
    
    async def _parse_csv_file(self, contents: bytes, filename: str, version_id: int) -> List[SheetInfo]:
        """Parse CSV file into a sheet"""
        sheet_create = SheetCreate(
            dataset_version_id=version_id,
            name=filename,
            sheet_index=0,
            description=None
        )

        sheet_id = await self.repository.create_sheet(sheet_create)

        # Create a BytesIO object from contents for pandas to read
        csv_buffer = BytesIO(contents)
        df = pd.read_csv(csv_buffer)

        # Create sheet metadata
        metadata = {
            "columns": len(df.columns),
            "rows": len(df)
            # Additional metadata can be added here
        }

        await self.repository.create_sheet_metadata(sheet_id, metadata)

        return [SheetInfo(
            id=sheet_id,
            name=filename,
            index=0,
            description=None
        )]
    
    async def _handle_unsupported_file_type(self, contents: bytes, filename: str, version_id: int) -> List[SheetInfo]:
        """Handle unsupported file types"""
        file_type = os.path.splitext(filename)[1].lower()[1:]
        sheet_create = SheetCreate(
            dataset_version_id=version_id,
            name=filename,
            sheet_index=0,
            description=f"Unsupported file type: {file_type}"
        )

        sheet_id = await self.repository.create_sheet(sheet_create)

        # Create minimal metadata
        metadata = {
            "file_type": file_type,
            "note": "Unsupported file type - no parsing performed"
        }

        await self.repository.create_sheet_metadata(sheet_id, metadata)

        return [SheetInfo(
            id=sheet_id,
            name=filename,
            index=0,
            description=f"Unsupported file type: {file_type}"
        )]
    
    async def _create_error_sheet(self, file: UploadFile, version_id: int, file_type: str, error_msg: str) -> List[SheetInfo]:
        """Create a sheet entry for a file parsing error"""
        filename = os.path.basename(file.filename)
        sheet_create = SheetCreate(
            dataset_version_id=version_id,
            name=filename,
            sheet_index=0,
            description="Error parsing file"
        )

        sheet_id = await self.repository.create_sheet(sheet_create)

        # Create error metadata
        metadata = {
            "error": error_msg,
            "file_type": file_type
        }

        await self.repository.create_sheet_metadata(sheet_id, metadata)

        return [SheetInfo(
            id=sheet_id,
            name=filename,
            index=0,
            description="Error parsing file"
        )]
    
    async def _generate_profile_report(self, data_source, sheet_name=None):
        """
        Generate profiling report for a dataset.
        This is a stub method to be implemented when profiling is needed.

        Args:
            data_source: The data source (Excel file, DataFrame, etc.)
            sheet_name: Optional sheet name for Excel files

        Returns:
            int: The file_id of the saved profiling report
        """
        # Implementation for profiling data goes here
        # For example, using pandas-profiling or other libraries
        # Create a report file and save it using self.repository.create_file
        # Return the file_id

        # Placeholder implementation
        return None

    # Dataset listing and retrieval methods
    async def list_datasets(
        self,
        limit: int = 10,
        offset: int = 0,
        sort_by: Optional[str] = None,
        sort_order: Optional[str] = None,
        name: Optional[str] = None,
        description: Optional[str] = None,
        created_by: Optional[int] = None,
        tags: Optional[List[str]] = None,
        file_type: Optional[str] = None,
        file_size_min: Optional[int] = None,
        file_size_max: Optional[int] = None,
        version_min: Optional[int] = None,
        version_max: Optional[int] = None,
        created_at_from: Optional[datetime] = None,
        created_at_to: Optional[datetime] = None,
        updated_at_from: Optional[datetime] = None,
        updated_at_to: Optional[datetime] = None
    ) -> List[Dataset]:
        """List datasets with optional filtering, sorting, and pagination"""
        # Validate and normalize inputs
        if limit < 1:
            limit = 10
        elif limit > 100:
            limit = 100

        if offset < 0:
            offset = 0

        # Default sort parameters
        valid_sort_fields = ["name", "created_at", "updated_at", "file_size", "current_version"]
        if sort_by not in valid_sort_fields:
            sort_by = "updated_at"

        valid_sort_orders = ["asc", "desc"]
        if sort_order not in valid_sort_orders:
            sort_order = "desc"
        
        # Call repository method
        result = await self.repository.list_datasets(
            limit=limit,
            offset=offset,
            sort_by=sort_by,
            sort_order=sort_order,
            name=name,
            description=description,
            created_by=created_by,
            tags=tags,
            file_type=file_type,
            file_size_min=file_size_min,
            file_size_max=file_size_max,
            version_min=version_min,
            version_max=version_max,
            created_at_from=created_at_from,
            created_at_to=created_at_to,
            updated_at_from=updated_at_from,
            updated_at_to=updated_at_to
        )

        # Transform raw results into appropriate response models
        # The repository now returns List[Dataset], so direct transformation might not be needed
        # or needs to be adjusted if the structure from repository.list_datasets is already Pydantic models.
        # Assuming result is List[Dataset] as per repository changes.
        return result

    async def get_dataset(self, dataset_id: int) -> Optional[Dataset]:
        """Get detailed information about a single dataset"""
        result = await self.repository.get_dataset(dataset_id)
        # No transformation needed if repository returns Optional[Dataset]
        return result

    async def update_dataset(self, dataset_id: int, data: DatasetUpdate) -> Optional[Dataset]:
        """Update dataset metadata including name, description, and tags"""
        # First check if dataset exists
        # get_dataset now returns Optional[Dataset]
        existing_dataset = await self.get_dataset(dataset_id)
        if not existing_dataset:
            return None

        # Update basic metadata
        updated_id = await self.repository.update_dataset(dataset_id, data)
        if not updated_id:
            return None

        # If tags were provided, update tags
        if data.tags is not None:
            # Delete all existing tags
            await self.repository.delete_dataset_tags(dataset_id)

            # Add new tags
            for tag_name in data.tags:
                tag_id = await self.repository.upsert_tag(tag_name)
                await self.repository.create_dataset_tag(dataset_id, tag_id)

        # Return the updated dataset
        return await self.get_dataset(dataset_id)

    async def list_dataset_versions(self, dataset_id: int) -> List[DatasetVersion]:
        """List all versions of a dataset"""
        # Check if dataset exists
        dataset = await self.get_dataset(dataset_id) # Returns Optional[Dataset]
        if not dataset:
            return [] # Return empty list if dataset not found

        result = await self.repository.list_dataset_versions(dataset_id)
        # No transformation needed if repository returns List[DatasetVersion]
        return result

    async def get_dataset_version(self, version_id: int) -> Optional[DatasetVersion]:
        """Get detailed information about a single dataset version"""
        result = await self.repository.get_dataset_version(version_id)
        # No transformation needed if repository returns Optional[DatasetVersion]
        # JSON parsing for sheets should ideally be handled within the repository or model itself if possible.
        # However, if sheets are part of the DatasetVersion Pydantic model and are complex,
        # the repository should ensure they are correctly populated.
        return result

    async def get_dataset_version_file(self, version_id: int) -> Optional[File]:
        """Get file information for a dataset version"""
        # Get the version first
        version = await self.get_dataset_version(version_id) # Returns Optional[DatasetVersion]
        if not version or not version.file_id: # Check if version and file_id exist
            return None

        # Get the associated file
        file_info = await self.repository.get_file(version.file_id) # Returns Optional[File]
        return file_info

    async def delete_dataset_version(self, version_id: int) -> bool:
        """Delete a dataset version"""
        # Check if version exists
        version = await self.get_dataset_version(version_id)
        if not version:
            return False

        deleted_id = await self.repository.delete_dataset_version(version_id)
        return deleted_id is not None

    async def list_tags(self) -> List[Tag]:
        """List all available tags"""
        result = await self.repository.list_tags()
        # No transformation needed if repository returns List[Tag]
        return result

    async def list_version_sheets(self, version_id: int) -> List[Sheet]:
        """Get all sheets for a dataset version"""
        # First check if version exists
        version = await self.get_dataset_version(version_id) # Returns Optional[DatasetVersion]
        if not version:
            return []

        # Get sheets
        sheets = await self.repository.list_version_sheets(version_id) # Returns List[Sheet]
        return sheets

    async def get_sheet_data(self, version_id: int, sheet_name: str, limit: int = 100, offset: int = 0) -> Tuple[List[str], List[Dict[str, Any]], bool]:
        """Get paginated data from a sheet"""
        # First check if version exists
        version = await self.get_dataset_version(version_id)
        if not version:
            return [], [], False

        # Get file info
        file_info = await self.repository.get_file(version.file_id)
        if not file_info or not file_info.file_data:
            return [], [], False

        # Get sheets to validate sheet_name
        sheets = await self.repository.list_version_sheets(version_id)
        sheet_names = [sheet.name for sheet in sheets]

        # For CSV, there should be just one sheet, use that
        if file_info.file_type == "csv" and not sheet_name and sheets:
            sheet_name = sheets[0].name

        # Validate sheet name
        if sheet_name not in sheet_names:
            return [], [], False

        # Get the data based on file type
        try:
            if file_info.file_type == "csv":
                return await self._get_csv_data(file_info.file_data, limit, offset)
            elif file_info.file_type in ["xls", "xlsx", "xlsm"]:
                return await self._get_excel_data(file_info.file_data, sheet_name, limit, offset)
            else:
                return [], [], False
        except Exception as e:
            print(f"Error reading file data: {str(e)}")
            return [], [], False

    async def _get_csv_data(self, file_data: bytes, limit: int, offset: int) -> Tuple[List[str], List[Dict[str, Any]], bool]:
        """Extract data from CSV file"""
        import csv
        import io

        # Create a file-like object from the bytes
        csv_buffer = io.StringIO(file_data.decode('utf-8', errors='replace'))

        # Create CSV reader
        reader = csv.reader(csv_buffer)

        # Read header row
        header = next(reader, [])
        if not header:
            return [], [], False

        # Skip rows for offset
        for _ in range(offset):
            if not next(reader, None):
                # Reached end of file
                return header, [], True

        # Read rows up to limit
        rows = []
        has_more = False
        for i, row in enumerate(reader):
            if i >= limit:
                has_more = True
                break

            # Convert to dictionary
            row_dict = {}
            for j, col in enumerate(header):
                if j < len(row):
                    row_dict[col] = row[j]
                else:
                    row_dict[col] = ""
            rows.append(row_dict)

        return header, rows, has_more

    async def _get_excel_data(self, file_data: bytes, sheet_name: str, limit: int, offset: int) -> Tuple[List[str], List[Dict[str, Any]], bool]:
        """Extract data from Excel file"""
        import openpyxl
        import io

        # Create a file-like object from the bytes
        excel_buffer = io.BytesIO(file_data)

        # Load workbook in read-only mode for better performance
        try:
            workbook = openpyxl.load_workbook(excel_buffer, read_only=True, data_only=True)
        except Exception as e:
            print(f"Error loading Excel file: {str(e)}")
            return [], [], False

        # Get sheet
        if sheet_name not in workbook.sheetnames:
            return [], [], False

        sheet = workbook[sheet_name]

        # Get header row
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], [], False

        header = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(rows[0])]

        # Skip rows for offset (add 1 to skip header)
        offset_index = offset + 1
        if offset_index >= len(rows):
            return header, [], False

        # Read rows up to limit
        data_rows = []
        for i in range(offset_index, min(offset_index + limit, len(rows))):
            row = rows[i]
            row_dict = {}
            for j, col in enumerate(header):
                if j < len(row):
                    row_dict[col] = row[j] if row[j] is not None else ""
                else:
                    row_dict[col] = ""
            data_rows.append(row_dict)

        has_more = (offset_index + limit) < len(rows)

        return header, data_rows, has_more

