"""DuckDB service for dataset operations - handles all file parsing and data operations using DuckDB"""
import duckdb
import os
import logging
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)

class DuckDBService:
    def __init__(self):
        self.conn = None
    
    def __enter__(self):
        self.conn = duckdb.connect(':memory:')
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.conn:
            self.conn.close()
    
    async def parse_excel_file(self, contents: bytes, filename: str) -> List[Dict[str, Any]]:
        """Parse Excel file using openpyxl and DuckDB for processing"""
        import openpyxl
        from io import BytesIO
        
        sheet_infos = []
        
        try:
            # Load workbook with openpyxl
            excel_buffer = BytesIO(contents)
            workbook = openpyxl.load_workbook(excel_buffer, read_only=True, data_only=True)
            
            for i, sheet_name in enumerate(workbook.sheetnames):
                try:
                    sheet = workbook[sheet_name]
                    
                    # Get basic metadata
                    # For read-only mode, we need to iterate to count rows
                    row_count = 0
                    col_count = 0
                    column_names = []
                    
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                        if row_idx == 0:  # First row - headers
                            column_names = [str(cell) if cell is not None else f"Column_{j+1}" 
                                          for j, cell in enumerate(row)]
                            col_count = len(column_names)
                        row_count += 1
                        
                        # Stop counting after a reasonable number to avoid performance issues
                        if row_count > 10000:
                            # Just estimate the rest
                            row_count = sheet.max_row if hasattr(sheet, 'max_row') else row_count
                            break
                    
                    sheet_infos.append({
                        'name': sheet_name,
                        'index': i,
                        'metadata': {
                            'columns': col_count,
                            'rows': row_count,
                            'column_names': column_names[:10]  # Store first 10 column names only
                        }
                    })
                except Exception as e:
                    logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                    sheet_infos.append({
                        'name': sheet_name,
                        'index': i,
                        'metadata': {
                            'error': str(e),
                            'columns': 0,
                            'rows': 0
                        }
                    })
            
            workbook.close()
                
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise
        
        return sheet_infos
    
    async def parse_csv_file(self, contents: bytes, filename: str) -> List[Dict[str, Any]]:
        """Parse CSV file using DuckDB and return sheet information"""
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp_file:
            tmp_file.write(contents)
            tmp_path = tmp_file.name
        
        try:
            with duckdb.connect(':memory:') as conn:
                # Read CSV file with error handling
                # First try with strict parsing to detect issues
                try:
                    conn.execute(f"CREATE TABLE csv_data AS SELECT * FROM read_csv_auto('{tmp_path}')")
                except Exception as e:
                    logger.warning(f"CSV parsing encountered errors, retrying with ignore_errors=true: {str(e)}")
                    # Retry with error ignoring
                    conn.execute(f"CREATE TABLE csv_data AS SELECT * FROM read_csv_auto('{tmp_path}', ignore_errors=true)")
                
                # Get metadata
                result = conn.execute("SELECT COUNT(*) as row_count FROM csv_data").fetchone()
                row_count = result[0] if result else 0
                
                # Get column info
                columns_result = conn.execute("DESCRIBE csv_data").fetchall()
                column_names = [col[0] for col in columns_result]
                col_count = len(column_names)
                
                return [{
                    'name': filename,
                    'index': 0,
                    'metadata': {
                        'columns': col_count,
                        'rows': row_count,
                        'column_names': column_names
                    }
                }]
        
        finally:
            # Clean up temp file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    async def parse_excel_file_from_info(self, file_info: Any, filename: str) -> List[Dict[str, Any]]:
        """Parse Excel file from file info object (supports both in-memory and file-based storage)"""
        if file_info.storage_type == "database":
            if file_info.file_data is None:
                raise ValueError(f"File data is missing for database-stored file with ID {file_info.id}")
            # Use in-memory data
            return await self.parse_excel_file(file_info.file_data, filename)
        elif file_info.storage_type == "filesystem" and hasattr(file_info, 'file_path'):
            # Read from filesystem
            with open(file_info.file_path, 'rb') as f:
                contents = f.read()
            return await self.parse_excel_file(contents, filename)
        else:
            raise ValueError(f"Unsupported storage type: {file_info.storage_type}")
    
    async def parse_csv_file_from_info(self, file_info: Any, filename: str) -> List[Dict[str, Any]]:
        """Parse CSV file from file info object (supports both in-memory and file-based storage)"""
        logger.debug(f"Parsing CSV file - Storage type: {file_info.storage_type}, "
                    f"Has file_data: {file_info.file_data is not None}, "
                    f"Has file_path: {hasattr(file_info, 'file_path')}")
        
        if file_info.storage_type == "database":
            if file_info.file_data is None:
                raise ValueError(f"File data is missing for database-stored file with ID {file_info.id}. "
                               f"This usually means the file was not properly saved to the database.")
            # Use in-memory data
            return await self.parse_csv_file(file_info.file_data, filename)
        elif file_info.storage_type == "filesystem":
            if not hasattr(file_info, 'file_path') or not file_info.file_path:
                raise ValueError(f"File path is missing for filesystem-stored file with ID {file_info.id}")
            # For filesystem storage, we can parse directly from file
            return await self._parse_csv_file_from_path(file_info.file_path, filename)
        else:
            raise ValueError(f"Unsupported storage type: {file_info.storage_type}. "
                           f"Expected 'database' or 'filesystem'.")
    
    async def _parse_csv_file_from_path(self, file_path: str, filename: str) -> List[Dict[str, Any]]:
        """Parse CSV file directly from filesystem path using DuckDB"""
        try:
            with duckdb.connect(':memory:') as conn:
                # Read CSV file directly from path with error handling
                conn.execute(f"CREATE TABLE csv_data AS SELECT * FROM read_csv_auto('{file_path}', ignore_errors=true)")
                
                # Get metadata
                result = conn.execute("SELECT COUNT(*) as row_count FROM csv_data").fetchone()
                row_count = result[0] if result else 0
                
                # Get column info
                columns_result = conn.execute("DESCRIBE csv_data").fetchall()
                column_names = [col[0] for col in columns_result]
                col_count = len(column_names)
                
                return [{
                    'name': filename,
                    'index': 0,
                    'metadata': {
                        'columns': col_count,
                        'rows': row_count,
                        'column_names': column_names
                    }
                }]
        except Exception as e:
            logger.error(f"Error parsing CSV file from path: {str(e)}")
            raise
    
    async def get_sheet_data_from_file_info(
        self,
        file_info: Any,
        sheet_name: Optional[str],
        limit: int,
        offset: int
    ) -> Tuple[List[str], List[Dict[str, Any]], bool]:
        """Get sheet data from file info object (supports both storage types)"""
        if file_info.file_type == "csv":
            if file_info.storage_type == "database":
                if file_info.file_data is None:
                    logger.error(f"File data is missing for database-stored CSV file with ID {file_info.id}")
                    return [], [], False
                return await self.get_sheet_data_from_csv(file_info.file_data, limit, offset)
            elif file_info.storage_type == "filesystem" and hasattr(file_info, 'file_path'):
                return await self._get_csv_data_from_path(file_info.file_path, limit, offset)
        elif file_info.file_type in ["xls", "xlsx", "xlsm"]:
            if file_info.storage_type == "database":
                if file_info.file_data is None:
                    logger.error(f"File data is missing for database-stored Excel file with ID {file_info.id}")
                    return [], [], False
                return await self.get_sheet_data_from_excel(file_info.file_data, sheet_name, limit, offset)
            elif file_info.storage_type == "filesystem" and hasattr(file_info, 'file_path'):
                with open(file_info.file_path, 'rb') as f:
                    file_data = f.read()
                return await self.get_sheet_data_from_excel(file_data, sheet_name, limit, offset)
        
        return [], [], False
    
    async def _get_csv_data_from_path(
        self,
        file_path: str,
        limit: int,
        offset: int
    ) -> Tuple[List[str], List[Dict[str, Any]], bool]:
        """Get paginated CSV data directly from filesystem path"""
        try:
            with duckdb.connect(':memory:') as conn:
                # Read CSV directly from path with error handling
                conn.execute(f"CREATE TABLE csv_data AS SELECT * FROM read_csv_auto('{file_path}', ignore_errors=true)")
                
                # Get column names
                columns_result = conn.execute("SELECT * FROM csv_data LIMIT 0").description
                headers = [col[0] for col in columns_result]
                
                # Get paginated data
                data_result = conn.execute(f"SELECT * FROM csv_data LIMIT {limit} OFFSET {offset}").fetchall()
                
                # Convert to list of dicts
                rows = []
                for row in data_result:
                    row_dict = {}
                    for i, value in enumerate(row):
                        row_dict[headers[i]] = value
                    rows.append(row_dict)
                
                # Check if there's more data
                count_result = conn.execute(f"SELECT COUNT(*) FROM csv_data WHERE ROWID > {offset + limit}").fetchone()
                has_more = count_result[0] > 0 if count_result else False
                
                return headers, rows, has_more
        except Exception as e:
            logger.error(f"Error reading CSV data from path: {str(e)}")
            return [], [], False
    
    async def get_sheet_data_from_excel(
        self, 
        file_data: bytes, 
        sheet_name: str, 
        limit: int, 
        offset: int
    ) -> Tuple[List[str], List[Dict[str, Any]], bool]:
        """Get paginated data from Excel sheet using openpyxl and DuckDB for processing"""
        import openpyxl
        from io import BytesIO
        
        try:
            # Load workbook with openpyxl
            excel_buffer = BytesIO(file_data)
            workbook = openpyxl.load_workbook(excel_buffer, read_only=True, data_only=True)
            
            if sheet_name not in workbook.sheetnames:
                return [], [], False
                
            sheet = workbook[sheet_name]
            
            # Read data into memory first
            all_data = []
            headers = None
            
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx == 0:  # Headers
                    headers = [str(cell) if cell is not None else f"Column_{i+1}" 
                             for i, cell in enumerate(row)]
                else:
                    # Convert row to dict
                    row_dict = {}
                    for i, value in enumerate(row[:len(headers)]):
                        row_dict[headers[i]] = value if value is not None else ""
                    all_data.append(row_dict)
            
            workbook.close()
            
            if not headers:
                return [], [], False
            
            # Now use DuckDB for efficient pagination
            with duckdb.connect(':memory:') as conn:
                # Create table from data
                if all_data:
                    # Create table with proper column types
                    create_cols = []
                    for col in headers:
                        create_cols.append(f'"{col}" VARCHAR')
                    
                    create_table_sql = f"CREATE TABLE sheet_data ({', '.join(create_cols)})"
                    conn.execute(create_table_sql)
                    
                    # Insert data
                    for row in all_data:
                        values = [str(row.get(col, '')) for col in headers]
                        placeholders = ', '.join(['?' for _ in values])
                        insert_sql = f"INSERT INTO sheet_data VALUES ({placeholders})"
                        conn.execute(insert_sql, values)
                    
                    # Get paginated data
                    data_result = conn.execute(f"SELECT * FROM sheet_data LIMIT {limit} OFFSET {offset}").fetchall()
                    
                    # Convert to list of dicts
                    rows = []
                    for row in data_result:
                        row_dict = {}
                        for i, value in enumerate(row):
                            row_dict[headers[i]] = value
                        rows.append(row_dict)
                    
                    # Check if there's more data
                    count_result = conn.execute(f"SELECT COUNT(*) FROM sheet_data WHERE ROWID > {offset + limit}").fetchone()
                    has_more = count_result[0] > 0 if count_result else False
                    
                    return headers, rows, has_more
                else:
                    return headers, [], False
                    
        except Exception as e:
            logger.error(f"Error reading Excel data: {str(e)}")
            return [], [], False
    
    async def get_sheet_data_from_csv(
        self, 
        file_data: bytes, 
        limit: int, 
        offset: int
    ) -> Tuple[List[str], List[Dict[str, Any]], bool]:
        """Get paginated data from CSV using DuckDB"""
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp_file:
            tmp_file.write(file_data)
            tmp_path = tmp_file.name
        
        try:
            with duckdb.connect(':memory:') as conn:
                # Read CSV with error handling
                conn.execute(f"CREATE TABLE csv_data AS SELECT * FROM read_csv_auto('{tmp_path}', ignore_errors=true)")
                
                # Get column names
                columns_result = conn.execute("SELECT * FROM csv_data LIMIT 0").description
                headers = [col[0] for col in columns_result]
                
                # Get paginated data
                data_result = conn.execute(f"SELECT * FROM csv_data LIMIT {limit} OFFSET {offset}").fetchall()
                
                # Convert to list of dicts
                rows = []
                for row in data_result:
                    row_dict = {}
                    for i, value in enumerate(row):
                        row_dict[headers[i]] = value
                    rows.append(row_dict)
                
                # Check if there's more data
                count_result = conn.execute(f"SELECT COUNT(*) FROM csv_data OFFSET {offset + limit}").fetchone()
                has_more = count_result[0] > 0 if count_result else False
                
                return headers, rows, has_more
        
        finally:
            # Clean up temp file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)