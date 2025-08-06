"""Data download API endpoints - True streaming implementation."""

from typing import Optional, AsyncIterator, List, Dict
from fastapi import APIRouter, Depends, Query, Path, HTTPException, status
from fastapi.responses import StreamingResponse
import json
import io
import csv
import logging
import zipfile
import asyncio
from concurrent.futures import ThreadPoolExecutor
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from ..api.models import CurrentUser
from ..core.authorization import get_current_user_info, require_dataset_read
from .dependencies import get_uow
from ..infrastructure.postgres.uow import PostgresUnitOfWork
from ..core.domain_exceptions import EntityNotFoundException

logger = logging.getLogger(__name__)


router = APIRouter(prefix="/datasets", tags=["downloads"])


async def _get_table_keys_for_commit(conn, commit_id: str) -> List[str]:
    """Get all unique table keys for a commit."""
    query = """
        SELECT DISTINCT 
            SPLIT_PART(logical_row_id, ':', 1) as table_key
        FROM dsa_core.commit_rows
        WHERE commit_id = $1
        ORDER BY table_key
    """
    
    rows = await conn.fetch(query, commit_id)
    return [row['table_key'] for row in rows]


def _get_raw_connection_from_uow(uow: PostgresUnitOfWork):
    """Pragmatic helper to get the raw asyncpg pool.
    
    TODO: A cleaner long-term solution would be to add a method to 
    PostgresUnitOfWork like `async def get_streaming_connection()` 
    that returns a raw connection for streaming use cases.
    """
    # This is a bit of a hack, but necessary if the UoW hides the pool.
    return uow._pool._pool if hasattr(uow._pool, '_pool') else uow._pool


def _parse_db_row(db_row: dict) -> dict:
    """Parse potentially nested and stringified JSON from the database."""
    row_data = db_row['data']
    if isinstance(row_data, str):
        row_data = json.loads(row_data)
    
    # Standardized data is nested under a 'data' key
    return row_data.get('data', row_data) if isinstance(row_data, dict) else row_data


async def _get_schema_headers(
    conn,
    commit_id: str,
    table_key: Optional[str] = None
) -> Optional[List[str]]:
    """Fetch the canonical column headers from the commit schema.
    
    Returns:
        List of column names in order, or None if schema not found
    """
    schema_row = await conn.fetchval(
        "SELECT schema_definition FROM dsa_core.commit_schemas WHERE commit_id = $1",
        commit_id
    )
    
    if not schema_row:
        return None
        
    schema_def = json.loads(schema_row) if isinstance(schema_row, str) else schema_row
    
    # If looking for a specific table
    if table_key and table_key in schema_def:
        columns = schema_def[table_key].get('columns', [])
        return [col['name'] for col in columns]
    
    # For full dataset download, we need to combine all tables' columns
    # or use the first table if there's only one
    if not table_key:
        all_columns = []
        for table_data in schema_def.values():
            if isinstance(table_data, dict) and 'columns' in table_data:
                for col in table_data['columns']:
                    if col['name'] not in all_columns:
                        all_columns.append(col['name'])
        return all_columns if all_columns else None
    
    return None


def _sanitize_column_name(name: str) -> str:
    """Sanitize column name for PostgreSQL identifier."""
    # Replace special characters with underscores
    sanitized = ''.join(c if c.isalnum() or c == '_' else '_' for c in name)
    # Ensure it doesn't start with a number
    if sanitized and sanitized[0].isdigit():
        sanitized = f"col_{sanitized}"
    # Ensure it's not empty
    if not sanitized:
        sanitized = "column"
    return sanitized.lower()


async def _stream_csv_direct_copy(
    conn,
    commit_id: str,
    table_key: Optional[str] = None,
    batch_size: int = 1000  # Process rows in batches
) -> AsyncIterator[bytes]:
    """
    Optimized CSV streaming with batching and better queries.
    
    Args:
        conn: Database connection
        commit_id: Commit to export from
        table_key: Optional table key to filter by
        batch_size: Number of rows to process in each batch
        
    Yields:
        CSV data chunks as bytes
    """
    # Simplified, index-friendly query
    if table_key:
        query = """
            SELECT r.data, cr.logical_row_id
            FROM dsa_core.commit_rows cr
            JOIN dsa_core.rows r ON cr.row_hash = r.row_hash
            WHERE cr.commit_id = $1 
            AND cr.logical_row_id LIKE $2
            ORDER BY cr.logical_row_id
        """
        params = [commit_id, f"{table_key}_%"]
    else:
        query = """
            SELECT r.data, cr.logical_row_id
            FROM dsa_core.commit_rows cr
            JOIN dsa_core.rows r ON cr.row_hash = r.row_hash
            WHERE cr.commit_id = $1
            ORDER BY cr.logical_row_id
        """
        params = [commit_id]
    
    # Get headers once
    headers = await _get_schema_headers(conn, commit_id, table_key)
    
    async with conn.transaction():
        cursor = conn.cursor(query, *params)
        
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        headers_written = False
        row_batch = []
        
        async for db_row in cursor:
            actual_data = _parse_db_row(db_row)
            
            if not headers_written:
                if not headers:
                    headers = list(actual_data.keys())
                writer.writerow(headers)
                yield output.getvalue().encode('utf-8')
                output.seek(0)
                output.truncate(0)
                headers_written = True
            
            # Batch rows instead of yielding each one
            row_batch.append([actual_data.get(h) for h in headers])
            
            if len(row_batch) >= batch_size:
                writer.writerows(row_batch)
                yield output.getvalue().encode('utf-8')
                output.seek(0)
                output.truncate(0)
                row_batch = []
        
        # Write remaining rows
        if row_batch:
            writer.writerows(row_batch)
            yield output.getvalue().encode('utf-8')
        
        if not headers_written:
            # If we have headers from schema but no data
            if headers:
                writer.writerow(headers)
                yield output.getvalue().encode('utf-8')
            else:
                yield b"No data found for this dataset\n"


async def _stream_parquet_data(
    conn,
    commit_id: str,
    table_key: Optional[str] = None
) -> bytes:
    """
    Generate Parquet data from database records.
    
    Args:
        conn: Database connection
        commit_id: Commit to export from
        table_key: Optional table key to filter by
        
    Returns:
        Parquet file as bytes
    """
    # Query for data
    if table_key:
        query = """
            SELECT r.data, cr.logical_row_id
            FROM dsa_core.commit_rows cr
            JOIN dsa_core.rows r ON cr.row_hash = r.row_hash
            WHERE cr.commit_id = $1 
            AND cr.logical_row_id LIKE $2
            ORDER BY cr.logical_row_id
        """
        params = [commit_id, f"{table_key}_%"]
    else:
        query = """
            SELECT r.data, cr.logical_row_id
            FROM dsa_core.commit_rows cr
            JOIN dsa_core.rows r ON cr.row_hash = r.row_hash
            WHERE cr.commit_id = $1
            ORDER BY cr.logical_row_id
        """
        params = [commit_id]
    
    # Get headers from schema
    headers = await _get_schema_headers(conn, commit_id, table_key)
    
    # Collect all data
    rows = []
    async with conn.transaction():
        cursor = conn.cursor(query, *params)
        async for db_row in cursor:
            actual_data = _parse_db_row(db_row)
            if not headers:
                headers = list(actual_data.keys())
            rows.append(actual_data)
    
    if not rows:
        # Create empty dataframe with headers
        if headers:
            data_dict = {h: [] for h in headers}
        else:
            data_dict = {"message": ["No data found"]}
        table = pa.Table.from_pydict(data_dict)
    else:
        # Convert to PyArrow table
        data_dict = {h: [] for h in headers}
        for row in rows:
            for h in headers:
                data_dict[h].append(row.get(h))
        
        table = pa.Table.from_pydict(data_dict)
    
    # Write to bytes buffer
    buffer = io.BytesIO()
    pq.write_table(table, buffer, compression='snappy')
    buffer.seek(0)
    return buffer.read()


async def _generate_excel_data(
    conn,
    commit_id: str,
    dataset_name: str
) -> bytes:
    """
    Generate Excel file with multiple sheets for all tables.
    
    Args:
        conn: Database connection
        commit_id: Commit to export from
        dataset_name: Name of the dataset for sheet naming
        
    Returns:
        Excel file as bytes
    """
    # Get all table keys
    table_keys = await _get_table_keys_for_commit(conn, commit_id)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    if not table_keys:
        # Create a sheet with no data message
        ws = wb.create_sheet(title="Empty")
        ws.append(["No data found for this dataset"])
    else:
        # Process each table
        for table_key in table_keys:
            # Create sheet with table name (Excel sheet names limited to 31 chars)
            sheet_name = table_key[:31] if len(table_key) > 31 else table_key
            ws = wb.create_sheet(title=sheet_name)
            
            # Query for data
            query = """
                SELECT r.data, cr.logical_row_id
                FROM dsa_core.commit_rows cr
                JOIN dsa_core.rows r ON cr.row_hash = r.row_hash
                WHERE cr.commit_id = $1 
                AND cr.logical_row_id LIKE $2
                ORDER BY cr.logical_row_id
            """
            params = [commit_id, f"{table_key}_%"]
            
            # Get headers from schema
            headers = await _get_schema_headers(conn, commit_id, table_key)
            
            # Collect data
            rows = []
            async with conn.transaction():
                cursor = conn.cursor(query, *params)
                async for db_row in cursor:
                    actual_data = _parse_db_row(db_row)
                    if not headers:
                        headers = list(actual_data.keys())
                    rows.append(actual_data)
            
            # Write headers with bold font
            if headers:
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                
                # Auto-adjust column widths based on header length
                for idx, header in enumerate(headers, 1):
                    col_letter = get_column_letter(idx)
                    ws.column_dimensions[col_letter].width = max(len(str(header)) + 2, 10)
            
            # Write data rows
            for row in rows:
                ws.append([row.get(h) for h in headers])
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


async def _create_streaming_download_response(
    dataset: dict,
    ref: dict,
    uow: PostgresUnitOfWork,
    table_key: Optional[str] = None,
    format: str = "csv"
) -> StreamingResponse:
    """Helper to create the streaming response for CSV or Parquet format.
    
    Note: Compression should be handled by FastAPI's GZipMiddleware at the
    application level, which will automatically compress responses when the
    client sends 'Accept-Encoding: gzip' header.
    """
    
    if format == "parquet":
        # Generate parquet data
        async def generate_parquet_stream():
            pool = _get_raw_connection_from_uow(uow)
            async with pool.acquire() as conn:
                parquet_data = await _stream_parquet_data(conn, ref['commit_id'], table_key)
                yield parquet_data
        
        filename = f"{dataset['name']}.parquet"
        if table_key:
            filename = f"{dataset['name']}_{table_key}.parquet"
        
        return StreamingResponse(
            generate_parquet_stream(),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    else:
        # Default to CSV
        async def generate_csv_stream():
            pool = _get_raw_connection_from_uow(uow)
            async with pool.acquire() as conn:
                async for chunk in _stream_csv_direct_copy(conn, ref['commit_id'], table_key):
                    yield chunk
        
        filename = f"{dataset['name']}.csv"
        if table_key:
            filename = f"{dataset['name']}_{table_key}.csv"
        
        return StreamingResponse(
            generate_csv_stream(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )


async def _create_streaming_zip_response(
    dataset: dict,
    ref: dict,
    uow: PostgresUnitOfWork,
    format: str = "csv"
) -> StreamingResponse:
    """Create a streaming ZIP response containing CSV or Parquet files for all tables."""
    
    async def generate_zip_stream():
        pool = _get_raw_connection_from_uow(uow)
        
        # We need to buffer the entire ZIP in memory since ZIP format requires
        # updating the central directory at the end
        zip_buffer = io.BytesIO()
        
        async with pool.acquire() as conn:
            # Get all table keys
            table_keys = await _get_table_keys_for_commit(conn, ref['commit_id'])
            
            if not table_keys:
                # If no tables, return an empty file in the ZIP
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    if format == "parquet":
                        # Create empty parquet file
                        empty_table = pa.Table.from_pydict({"message": ["No data found"]})
                        empty_buffer = io.BytesIO()
                        pq.write_table(empty_table, empty_buffer, compression='snappy')
                        zip_file.writestr(f"{dataset['name']}_empty.parquet", empty_buffer.getvalue())
                    else:
                        zip_file.writestr(f"{dataset['name']}_empty.csv", "No data found for this dataset\n")
                zip_buffer.seek(0)
                yield zip_buffer.read()
                return
            
            # Create ZIP file
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for table_key in table_keys:
                    if format == "parquet":
                        # Generate parquet data for this table
                        parquet_data = await _stream_parquet_data(conn, ref['commit_id'], table_key)
                        filename = f"{dataset['name']}_{table_key}.parquet"
                        zip_file.writestr(filename, parquet_data)
                    else:
                        # Collect CSV data for this table
                        csv_data = io.BytesIO()
                        async for chunk in _stream_csv_direct_copy(conn, ref['commit_id'], table_key):
                            csv_data.write(chunk)
                        
                        # Add to ZIP with table-specific filename
                        csv_filename = f"{dataset['name']}_{table_key}.csv"
                        zip_file.writestr(csv_filename, csv_data.getvalue())
            
            # Stream the complete ZIP file
            zip_buffer.seek(0)
            # Stream in chunks to avoid loading entire file in memory at once
            chunk_size = 1024 * 1024  # 1MB chunks
            while True:
                chunk = zip_buffer.read(chunk_size)
                if not chunk:
                    break
                yield chunk
    
    filename = f"{dataset['name']}.zip"
    
    return StreamingResponse(
        generate_zip_stream(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{dataset_id}/refs/{ref_name}/tables/{table_key}/download")
async def download_table(
    dataset_id: int = Path(..., description="Dataset ID"),
    ref_name: str = Path(..., description="Ref/branch name"),
    table_key: str = Path(..., description="Table key"),
    format: str = Query("csv", description="Export format (csv or parquet)"),
    columns: Optional[str] = Query(None, description="Column selection not yet supported"),
    current_user: CurrentUser = Depends(get_current_user_info),
    uow: PostgresUnitOfWork = Depends(get_uow),
    _: CurrentUser = Depends(require_dataset_read)
):
    """Download specific table in CSV or Parquet format."""
    if format not in ["csv", "parquet"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Format '{format}' not supported. Use 'csv' or 'parquet'."
        )
    
    if columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Column selection not yet supported in streaming implementation"
        )
    
    # Get dataset and ref
    dataset = await uow.datasets.get_dataset_by_id(dataset_id)
    if not dataset:
        raise EntityNotFoundException("Dataset", dataset_id)
    
    ref = await uow.commits.get_ref(dataset_id, ref_name)
    if not ref:
        raise EntityNotFoundException("Ref", ref_name)
    
    return await _create_streaming_download_response(dataset, ref, uow, table_key, format)


@router.get("/{dataset_id}/refs/{ref_name}/download")
async def download_dataset(
    dataset_id: int = Path(..., description="Dataset ID"),
    ref_name: str = Path(..., description="Ref/branch name"),
    format: str = Query("zip", description="Export format (zip, zip-parquet, excel, csv, or parquet)"),
    current_user: CurrentUser = Depends(get_current_user_info),
    uow: PostgresUnitOfWork = Depends(get_uow),
    _: CurrentUser = Depends(require_dataset_read)
):
    """Download entire dataset as ZIP file, Excel file, or single file format."""
    if format not in ["csv", "zip", "parquet", "zip-parquet", "excel"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Format '{format}' not supported. Use 'zip' (CSV files), 'zip-parquet' (Parquet files), 'excel' (Excel with sheets), or 'csv'/'parquet' for single-table datasets."
        )
    
    # Get dataset and ref
    dataset = await uow.datasets.get_dataset_by_id(dataset_id)
    if not dataset:
        raise EntityNotFoundException("Dataset", dataset_id)
    
    ref = await uow.commits.get_ref(dataset_id, ref_name)
    if not ref:
        raise EntityNotFoundException("Ref", ref_name)
    
    if format == "excel":
        # Return Excel file with multiple sheets
        async def generate_excel_stream():
            pool = _get_raw_connection_from_uow(uow)
            async with pool.acquire() as conn:
                excel_data = await _generate_excel_data(conn, ref['commit_id'], dataset['name'])
                yield excel_data
        
        filename = f"{dataset['name']}.xlsx"
        return StreamingResponse(
            generate_excel_stream(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    elif format == "zip":
        # Return ZIP with CSV files
        return await _create_streaming_zip_response(dataset, ref, uow, "csv")
    elif format == "zip-parquet":
        # Return ZIP with Parquet files
        return await _create_streaming_zip_response(dataset, ref, uow, "parquet")
    else:
        # Single file format (csv or parquet): Check if single table
        pool = _get_raw_connection_from_uow(uow)
        async with pool.acquire() as conn:
            table_keys = await _get_table_keys_for_commit(conn, ref['commit_id'])
            
            if len(table_keys) > 1:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Dataset has {len(table_keys)} tables. Use format=zip, format=zip-parquet, or format=excel for multi-table datasets."
                )
            elif len(table_keys) == 1:
                # Single table - can return as single file
                return await _create_streaming_download_response(dataset, ref, uow, table_keys[0], format)
            else:
                # No tables
                return await _create_streaming_download_response(dataset, ref, uow, None, format)

