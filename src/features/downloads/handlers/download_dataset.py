"""Handler for downloading datasets in various formats."""

import io
import csv
import json
from typing import Optional, Dict, Any, List
from dataclasses import dataclass
import openpyxl
from openpyxl import Workbook

from src.core.abstractions import IUnitOfWork, ITableMetadataReader, ITableDataReader
from src.core.domain_exceptions import EntityNotFoundException, ValidationException


@dataclass
class DownloadDatasetCommand:
    user_id: int
    dataset_id: int
    ref_name: str
    format: str = "csv"  # csv, excel, json
    table_key: Optional[str] = None  # If None, download all tables


class DownloadDatasetHandler:
    """Handler for downloading dataset data."""
    
    def __init__(self, uow: IUnitOfWork, metadata_reader: ITableMetadataReader, data_reader: ITableDataReader):
        self._uow = uow
        self._metadata_reader = metadata_reader
        self._data_reader = data_reader
    
    async def handle(self, command: DownloadDatasetCommand) -> Dict[str, Any]:
        """
        Download dataset data in specified format.
        
        Returns dict with:
        - content: file content (bytes or string)
        - content_type: MIME type
        - filename: suggested filename
        """
        # Validate format
        if command.format not in ["csv", "excel", "json"]:
            raise ValidationException(f"Unsupported format: {command.format}", field="format")
        
        # Get dataset info
        dataset = await self._uow.datasets.get_dataset_by_id(command.dataset_id)
        if not dataset:
            raise EntityNotFoundException("Dataset", command.dataset_id)
        
        # Get ref
        ref = await self._uow.commits.get_ref(command.dataset_id, command.ref_name)
        if not ref:
            raise EntityNotFoundException("Ref", command.ref_name)
        
        commit_id = ref['commit_id']
        
        # Get data
        if command.table_key:
            # Single table
            data = await self._get_all_table_data(commit_id, command.table_key)
            columns = await self._get_table_columns(commit_id, command.table_key, data)
        else:
            # All tables - for CSV/Excel we need to handle multiple tables
            if command.format == "excel":
                return await self._format_all_tables_excel(commit_id, dataset['name'])
            elif command.format == "json":
                return await self._format_all_tables_json(commit_id, dataset['name'])
            else:  # CSV - default to primary table
                data = await self._get_all_table_data(commit_id, "primary")
                columns = await self._get_table_columns(commit_id, "primary", data)
        
        # Format data based on requested format
        if command.format == "csv":
            return self._format_as_csv(data, columns, dataset['name'])
        elif command.format == "excel":
            return self._format_as_excel(data, columns, dataset['name'])
        else:  # json
            return self._format_as_json(data, dataset['name'])
    
    def _format_as_csv(self, data: List[Dict], columns: List[str], dataset_name: str) -> Dict[str, Any]:
        """Format data as CSV."""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        writer.writerows(data)
        
        return {
            "content": output.getvalue().encode('utf-8'),
            "content_type": "text/csv",
            "filename": f"{dataset_name}.csv"
        }
    
    def _format_as_excel(self, data: List[Dict], columns: List[str], dataset_name: str) -> Dict[str, Any]:
        """Format data as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "content": output.getvalue(),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": f"{dataset_name}.xlsx"
        }
    
    def _format_as_json(self, data: List[Dict], dataset_name: str) -> Dict[str, Any]:
        """Format data as JSON."""
        json_str = json.dumps(data, indent=2, default=str)
        
        return {
            "content": json_str.encode('utf-8'),
            "content_type": "application/json",
            "filename": f"{dataset_name}.json"
        }
    
    async def _get_all_table_data(self, commit_id: str, table_key: str) -> List[Dict]:
        """Get all data for a table in batches."""
        all_data = []
        offset = 0
        batch_size = 1000
        
        while True:
            result = await self._data_reader.get_table_data(
                commit_id=commit_id,
                table_key=table_key,
                offset=offset,
                limit=batch_size
            )
            
            all_data.extend(result)
            offset += batch_size
            
            if len(result) < batch_size:
                break
        
        return all_data
    
    async def _get_table_columns(self, commit_id: str, table_key: str, data: List[Dict]) -> List[str]:
        """Get column names for a table."""
        # Try to get from schema first
        schema = await self._metadata_reader.get_table_schema(commit_id, table_key)
        if schema and 'columns' in schema:
            return [col['name'] for col in schema['columns']]
        
        # Otherwise get from data
        if data:
            return list(data[0].keys())
        
        return []
    
    async def _format_all_tables_excel(self, commit_id: str, dataset_name: str) -> Dict[str, Any]:
        """Format all tables as Excel file."""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Get all table keys
        table_keys = await self._metadata_reader.list_table_keys(commit_id)
        if not table_keys:
            table_keys = ['primary']  # Default to primary
        
        for table_key in table_keys:
            # Create sheet for each table
            ws = wb.create_sheet(title=table_key[:31])  # Excel sheet name limit
            
            # Get data
            data = await self._get_all_table_data(commit_id, table_key)
            
            if data:
                # Write headers
                headers = list(data[0].keys())
                for col_idx, col_name in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                
                # Write data
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, col_name in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            "content": output.getvalue(),
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": f"{dataset_name}.xlsx"
        }
    
    async def _format_all_tables_json(self, commit_id: str, dataset_name: str) -> Dict[str, Any]:
        """Format all tables as JSON."""
        result = {
            "dataset_name": dataset_name,
            "commit_id": commit_id,
            "tables": {}
        }
        
        # Get all table keys
        table_keys = await self._metadata_reader.list_table_keys(commit_id)
        if not table_keys:
            table_keys = ['primary']  # Default to primary
        
        for table_key in table_keys:
            # Get data
            data = await self._get_all_table_data(commit_id, table_key)
            
            # Get schema
            schema = await self._metadata_reader.get_table_schema(commit_id, table_key)
            
            result["tables"][table_key] = {
                "schema": schema,
                "row_count": len(data),
                "data": data
            }
        
        json_str = json.dumps(result, indent=2, default=str)
        
        return {
            "content": json_str.encode('utf-8'),
            "content_type": "application/json",
            "filename": f"{dataset_name}.json"
        }